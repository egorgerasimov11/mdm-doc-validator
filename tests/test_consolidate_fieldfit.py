"""fieldfit: fit values into SAP field lengths so SAP doesn't warn 'Data loss'.
Long names/street spill into continuation fields (data preserved); search keys /
legacy fields truncate; SOURCE_ADDRNUMBER blanks; bank/tax never touched."""
from __future__ import annotations

import copy
import io

import openpyxl

from consolidation_helpers import make_americas_form, make_template, needs_converter
from mdmdoc.consolidation import fieldfit
from mdmdoc.consolidation.template_io import BPTemplate

LONG_NAME = "SHANGHAI FOREIGN SERVICE (BEIJING) CO., LTD."          # 44 chars
LONG_STREET = ("ROOM 416, 4TH FLOOR, BUILDING 2, NO. 1A AND NO. 3, "
               "GUANGHUA ROAD CHAOYANG DISTRICT")                   # > 60 chars


def test_name_spills_into_continuation_preserving_data():
    rows = {"LFA1 - Supplier General": [{"NAME1": LONG_NAME}]}
    fieldfit.fit_sap_fields(rows)
    r = rows["LFA1 - Supplier General"][0]
    assert len(r["NAME1"]) <= 35
    assert r.get("NAME2")                                   # overflow spilled
    assert " ".join((r["NAME1"] + " " + r["NAME2"]).split()) == \
           " ".join(LONG_NAME.split())                      # full name recoverable


def test_name_splits_identically_across_sheets():
    # the vendor name is a cross-sheet invariant → NAME1 must match on every sheet
    rows = {
        "LFA1 - Supplier General": [{"NAME1": LONG_NAME}],
        "ADRC - Address": [{"NAME1": LONG_NAME}],
        "BUT000 - General": [{"NAME_ORG1": LONG_NAME}],
    }
    fieldfit.fit_sap_fields(rows)
    n_lfa1 = rows["LFA1 - Supplier General"][0]["NAME1"]
    assert n_lfa1 == rows["ADRC - Address"][0]["NAME1"]
    assert n_lfa1 == rows["BUT000 - General"][0]["NAME_ORG1"]
    assert len(n_lfa1) <= 35


def test_street_spills_into_str_suppl():
    rows = {"ADRC - Address": [{"STREET": LONG_STREET}]}
    fieldfit.fit_sap_fields(rows)
    r = rows["ADRC - Address"][0]
    assert len(r["STREET"]) <= 35    # SAP restricts STREET chars 36-60 → fit to 35
    assert r.get("STR_SUPPL1")                              # overflow spilled


def test_search_and_legacy_fields_truncated():
    rows = {
        "LFA1 - Supplier General": [{"SORTL": LONG_NAME, "STRAS": LONG_STREET}],
        "BUT000 - General": [{"BU_SORT1": LONG_NAME}],
    }
    fieldfit.fit_sap_fields(rows)
    assert len(rows["LFA1 - Supplier General"][0]["SORTL"]) == 10
    assert len(rows["LFA1 - Supplier General"][0]["STRAS"]) == 35
    assert len(rows["BUT000 - General"][0]["BU_SORT1"]) == 20


def test_bank_and_tax_fields_never_truncated():
    long_val = "1" * 40
    rows = {
        "BUT0BK - Bank Account": [{"BANKN": long_val, "IBAN": long_val,
                                   "BANKL": long_val}],
        "DFKKBPTAXNUM - Tax Number": [{"TAXNUM": long_val}],
        "LFA1 - Supplier General": [{"STCD1": long_val}],
    }
    fieldfit.fit_sap_fields(rows)
    assert rows["BUT0BK - Bank Account"][0]["BANKN"] == long_val
    assert rows["BUT0BK - Bank Account"][0]["IBAN"] == long_val
    assert rows["DFKKBPTAXNUM - Tax Number"][0]["TAXNUM"] == long_val
    assert rows["LFA1 - Supplier General"][0]["STCD1"] == long_val


def test_deterministic():
    base = {"LFA1 - Supplier General": [{"NAME1": LONG_NAME, "SORTL": LONG_NAME}]}
    assert fieldfit.fit_sap_fields(copy.deepcopy(base)) == \
           fieldfit.fit_sap_fields(copy.deepcopy(base))


def test_spill_only_into_existing_columns(tmp_path):
    # a template lacking NAME2 must not gain a NAME2 the writer would reject
    sheets = {"LFA1 - Supplier General": ["_COMMENT", "SOURCE_ID", "NAME1"]}
    tpl = BPTemplate(make_template(tmp_path / "t.xlsx", sheets))
    rows = {"LFA1 - Supplier General": [{"NAME1": LONG_NAME}]}
    fieldfit.fit_sap_fields(rows, tpl)
    r = rows["LFA1 - Supplier General"][0]
    assert len(r["NAME1"]) <= 35
    assert "NAME2" not in r                                  # no column → don't add
    tpl.close()


@needs_converter
def test_consolidate_spills_long_name_e2e(monkeypatch, tmp_path):
    from fastapi.testclient import TestClient
    from mdmdoc import config, runstore
    from mdmdoc.server.app import create_app
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)
    monkeypatch.setenv("MDMDOC_BANK_VALUES", "full")
    client = TestClient(create_app("full"))
    form = make_americas_form(tmp_path / "f.xlsm", vendor_name=LONG_NAME)
    tpl = make_template(tmp_path / "t.xlsx")
    case = client.post("/ui/consolidation/new", follow_redirects=False
                       ).headers["location"].rsplit("/", 1)[-1]
    client.post(f"/ui/consolidation/{case}/template",
                files={"file": ("t.xlsx", io.BytesIO(tpl.read_bytes()))})
    client.post(f"/ui/consolidation/{case}/extract",
                files={"file": ("f.xlsm", io.BytesIO(form.read_bytes()))})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    dl = client.get(f"/ui/consolidation/{case}/download")
    ws = openpyxl.load_workbook(io.BytesIO(dl.content))["LFA1 - Supplier General"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    spilled = False
    for r_ in range(3, ws.max_row + 1):
        n1 = ws.cell(row=r_, column=h["NAME1"]).value
        if n1:
            assert len(str(n1)) <= 35
            if ws.cell(row=r_, column=h["NAME2"]).value:
                spilled = True
    assert spilled                                          # long name spilled into NAME2
