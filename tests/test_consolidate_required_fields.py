"""required_fields: SAP-mandatory KEY fields (DATE_FROM …) filled from defaults."""
from __future__ import annotations

import openpyxl

from consolidation_helpers import make_template
from mdmdoc.consolidation import required_fields
from mdmdoc.consolidation.template_io import BPTemplate


def _template_with_address_keys(tmp_path):
    """A mini template whose ADR2/ADRC headers mark DATE_FROM/NATION/CONSNUMBER
    as key fields via underline (as the real SAP template does)."""
    from consolidation_helpers import TEMPLATE_SHEETS
    sheets = {k: list(v) for k, v in TEMPLATE_SHEETS.items()}
    # EXTEND the real converter columns with the address key fields
    sheets["ADR2 - Phone"] += ["DATE_FROM", "CONSNUMBER"]
    sheets["ADRC - Address"] += ["DATE_FROM", "NATION"]
    path = make_template(tmp_path / "t.xlsx", sheets)
    # underline the key headers
    wb = openpyxl.load_workbook(path)
    from openpyxl.styles import Font
    keymap = {"ADR2 - Phone": {"SOURCE_ID", "SOURCE_ADDRNUMBER", "DATE_FROM", "CONSNUMBER"},
              "ADRC - Address": {"SOURCE_ID", "SOURCE_ADDRNUMBER", "DATE_FROM", "NATION"}}
    for sheet, keys in keymap.items():
        ws = wb[sheet]
        for c in ws[2]:
            if c.value in keys:
                c.font = Font(underline="single")
    wb.save(path)
    wb.close()
    return path


def test_key_fields_read_from_underline(tmp_path):
    tpl = BPTemplate(_template_with_address_keys(tmp_path))
    kf = tpl.key_fields()
    assert kf["ADR2 - Phone"] == {"SOURCE_ID", "SOURCE_ADDRNUMBER", "DATE_FROM", "CONSNUMBER"}
    assert "DATE_FROM" in kf["ADRC - Address"] and "NATION" in kf["ADRC - Address"]
    tpl.close()


def test_fill_dates_and_consnumber_not_nation(tmp_path):
    tpl = BPTemplate(_template_with_address_keys(tmp_path))
    kf = tpl.key_fields()
    rows = {
        "ADR2 - Phone": [{"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "N1",
                          "TEL_NUMBER": "010-1"}],
        "ADRC - Address": [{"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "N1",
                            "NAME1": "X"}],
    }
    required_fields.fill(rows, kf, "20260711")
    adr2 = rows["ADR2 - Phone"][0]
    assert adr2["DATE_FROM"] == "20260711"
    assert adr2["CONSNUMBER"] == "001"
    adrc = rows["ADRC - Address"][0]
    assert adrc["DATE_FROM"] == "20260711"
    assert "NATION" not in adrc or not adrc["NATION"]   # NATION stays blank
    tpl.close()


def test_fill_does_not_overwrite_present_values(tmp_path):
    tpl = BPTemplate(_template_with_address_keys(tmp_path))
    kf = tpl.key_fields()
    rows = {"ADR2 - Phone": [{"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "N1",
                             "DATE_FROM": "00010101", "TEL_NUMBER": "1"}]}
    required_fields.fill(rows, kf, "20260711")
    assert rows["ADR2 - Phone"][0]["DATE_FROM"] == "00010101"  # kept


def test_apply_constants_overrides_assignment_id():
    rows = {
        "LFB1 - Company Code (Supplier)": [{"SOURCE_ID": "N1",
                                            "ASSIGNMENT_ID": "N1"}],
        "LFM1 - Purchasing Org Data": [{"ASSIGNMENT_ID": "NEW_20260711_01"}],
        "LFA1 - Supplier General": [{"SOURCE_ID": "N1", "NAME1": "X"}],
    }
    required_fields.apply_constants(rows)
    assert rows["LFB1 - Company Code (Supplier)"][0]["ASSIGNMENT_ID"] == "000000000001"
    assert rows["LFM1 - Purchasing Org Data"][0]["ASSIGNMENT_ID"] == "000000000001"
    # a sheet with no ASSIGNMENT_ID column is left untouched
    assert "ASSIGNMENT_ID" not in rows["LFA1 - Supplier General"][0]


def test_apply_constants_sets_bp_category_only_on_but000():
    rows = {
        "BUT000 - General": [{"NAME_ORG1": "X"}],
        "BUT0ID - Identifier": [{"TYPE": "SAPGLOBAL_VATR", "IDNUMBER": "1"}],
        "DFKKBPTAXNUM - Tax Number": [{"TAXTYPE": "CN0", "TAXNUM": "9"}],
    }
    required_fields.apply_constants(rows)
    assert rows["BUT000 - General"][0]["TYPE"] == "2"                   # BP category
    assert rows["BUT0ID - Identifier"][0]["TYPE"] == "SAPGLOBAL_VATR"   # untouched
    assert rows["DFKKBPTAXNUM - Tax Number"][0]["TAXTYPE"] == "CN0"     # untouched


def test_apply_constants_sets_language_and_reprf():
    rows = {
        "LFA1 - Supplier General": [{"SPRAS": "ZH"}],
        "ADRC - Address": [{"LANGU": "ZH"}, {"LANGU": "ZH"}],   # both version rows
        "BUT000 - General": [{"BU_LANGU": "ZH", "LANGU_CORR": "ZH"}],
        "LFB1 - Company Code (Supplier)": [{"BUKRS": "0497"}],
    }
    required_fields.apply_constants(rows)                       # columns=None → set all
    # address + vendor language = the SAP 1-char key "E"
    assert rows["LFA1 - Supplier General"][0]["SPRAS"] == "E"
    assert [r["LANGU"] for r in rows["ADRC - Address"]] == ["E", "E"]
    # BUT000 BP-level language blanked (person-only on an org)
    assert rows["BUT000 - General"][0]["BU_LANGU"] == ""
    assert rows["BUT000 - General"][0]["LANGU_CORR"] == ""
    assert rows["LFB1 - Company Code (Supplier)"][0]["REPRF"] == "X"


def test_consolidate_sets_assignment_id_constant(monkeypatch, tmp_path):
    # end-to-end: LFB1 rows carry the vendor assignment constant, not SOURCE_ID
    import io
    from fastapi.testclient import TestClient
    from mdmdoc import config, runstore
    from mdmdoc.server.app import create_app
    from consolidation_helpers import make_americas_form
    if not __import__("mdmdoc.consolidation", fromlist=["available"]).available():
        import pytest
        pytest.skip("converter not installed")
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)
    monkeypatch.setenv("MDMDOC_BANK_VALUES", "full")
    client = TestClient(create_app("full"))
    form = make_americas_form(tmp_path / "f.xlsm")
    tpl = _template_with_address_keys(tmp_path)
    case = client.post("/ui/consolidation/new", follow_redirects=False
                       ).headers["location"].rsplit("/", 1)[-1]
    client.post(f"/ui/consolidation/{case}/template",
                files={"file": ("t.xlsx", io.BytesIO(tpl.read_bytes()))})
    client.post(f"/ui/consolidation/{case}/extract",
                files={"file": ("f.xlsm", io.BytesIO(form.read_bytes()))})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    dl = client.get(f"/ui/consolidation/{case}/download")
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["LFB1 - Company Code (Supplier)"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    vals = [ws.cell(row=r_, column=h["ASSIGNMENT_ID"]).value
            for r_ in range(3, ws.max_row + 1)]
    filled = [v for v in vals if v]
    assert filled and all(v == "000000000001" for v in filled), vals
    wb.close()


def test_consolidate_writes_date_from(monkeypatch, tmp_path):
    # end-to-end via the API: the ADR-family rows land with DATE_FROM filled
    import io
    from fastapi.testclient import TestClient
    from mdmdoc import config, runstore
    from mdmdoc.server.app import create_app
    from consolidation_helpers import make_americas_form, needs_converter
    if not __import__("mdmdoc.consolidation", fromlist=["available"]).available():
        import pytest
        pytest.skip("converter not installed")
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)
    monkeypatch.setenv("MDMDOC_BANK_VALUES", "full")
    client = TestClient(create_app("full"))
    form = make_americas_form(tmp_path / "f.xlsm")  # has a phone? SETH form has none, but ADRC yes
    tpl = _template_with_address_keys(tmp_path)
    case = client.post("/ui/consolidation/new", follow_redirects=False
                       ).headers["location"].rsplit("/", 1)[-1]
    client.post(f"/ui/consolidation/{case}/template",
                files={"file": ("t.xlsx", io.BytesIO(tpl.read_bytes()))})
    client.post(f"/ui/consolidation/{case}/extract",
                files={"file": ("f.xlsm", io.BytesIO(form.read_bytes()))})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    dl = client.get(f"/ui/consolidation/{case}/download")
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["ADRC - Address"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    # ADRC row exists (SETH form has an address) with today's DATE_FROM
    today = __import__("datetime").datetime.now().strftime("%Y%m%d")
    vals = [ws.cell(row=r_, column=h["DATE_FROM"]).value for r_ in range(3, ws.max_row + 1)]
    assert today in vals
    wb.close()
