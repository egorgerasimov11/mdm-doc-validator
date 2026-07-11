"""SAP-import compatibility of the output workbook: shared strings (not
openpyxl inline), customer sheets hidden, structure well-formed."""
from __future__ import annotations

import xml.dom.minidom as minidom
import zipfile

import openpyxl

from consolidation_helpers import make_template, prefill_vendor
from mdmdoc.consolidation.template_io import BPTemplate

ROWS = {
    "LFA1 - Supplier General": [{"SOURCE_ID": "N1", "NAME1": "上海外服 LLC", "LAND1": "CN"}],
    "ADR2 - Phone": [{"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "N1",
                      "TEL_NUMBER": "010-56063579"}],
}


def _out(tmp_path):
    tpl = BPTemplate(make_template(tmp_path / "t.xlsx"))
    tpl.append_rows(ROWS)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    return out


def test_output_uses_shared_strings_not_inline(tmp_path):
    out = _out(tmp_path)
    z = zipfile.ZipFile(out)
    assert "xl/sharedStrings.xml" in z.namelist()
    inline = sum(z.read(n).count(b"inlineStr")
                 for n in z.namelist() if n.startswith("xl/worksheets/"))
    assert inline == 0
    shared = sum(z.read(n).count(b't="s"')
                 for n in z.namelist() if n.startswith("xl/worksheets/"))
    assert shared > 0
    # Content_Types + rels reference sharedStrings
    assert b"sharedStrings.xml" in z.read("[Content_Types].xml")
    assert b"sharedStrings.xml" in z.read("xl/_rels/workbook.xml.rels")


def test_output_parts_well_formed(tmp_path):
    out = _out(tmp_path)
    z = zipfile.ZipFile(out)
    for n in ("[Content_Types].xml", "xl/sharedStrings.xml",
              "xl/_rels/workbook.xml.rels"):
        minidom.parseString(z.read(n))  # raises on malformed
    for n in z.namelist():
        if n.startswith("xl/worksheets/") and n.endswith(".xml"):
            minidom.parseString(z.read(n))


def test_output_reopens_and_values_intact(tmp_path):
    out = _out(tmp_path)
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["LFA1 - Supplier General"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    got = [ws.cell(row=r, column=h["NAME1"]).value for r in range(3, ws.max_row + 1)]
    assert "上海外服 LLC" in got
    wb.close()


def test_customer_sheets_hidden_vendor_visible(tmp_path):
    # give the mini-template a customer sheet so there is one to hide
    from consolidation_helpers import TEMPLATE_SHEETS, make_template as mk
    sheets = dict(TEMPLATE_SHEETS)
    sheets["KNA1 - Customer General"] = ["_COMMENT", "SOURCE_ID", "KUNNR", "NAME1"]
    path = mk(tmp_path / "twc.xlsx", sheets)
    tpl = BPTemplate(path)
    tpl.append_rows(ROWS)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    wb = openpyxl.load_workbook(out)
    assert wb["KNA1 - Customer General"].sheet_state == "hidden"
    assert wb["LFA1 - Supplier General"].sheet_state == "visible"
    wb.close()


def test_empty_string_cells_do_not_break(tmp_path):
    # a prefilled + appended template that also has blank cells round-trips
    path = prefill_vendor(make_template(tmp_path / "t.xlsx"))
    tpl = BPTemplate(path)
    tpl.append_rows(ROWS)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    z = zipfile.ZipFile(out)
    inline = sum(z.read(n).count(b"inlineStr")
                 for n in z.namelist() if n.startswith("xl/worksheets/"))
    assert inline == 0
    openpyxl.load_workbook(out).close()  # still opens
