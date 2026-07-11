"""verify.py tripwires — deliberately corrupt the output between write and
verify and assert the verifier screams. The heart of «несколько раз проверил»."""
from __future__ import annotations

import openpyxl
import pytest

from consolidation_helpers import make_template, prefill_vendor
from mdmdoc.consolidation.template_io import BPTemplate
from mdmdoc.consolidation.verify import verify_output

SID = "NEW_20260710_01"
ROWS = {
    "LFA1 - Supplier General": [
        {"SOURCE_ID": SID, "NAME1": "ALPHA LLC", "LAND1": "US", "ORT01": "DENVER"}],
    "BUT000 - General": [{"SOURCE_ID": SID, "NAME_ORG1": "ALPHA LLC"}],
    "ADRC - Address": [{"SOURCE_ID": SID, "SOURCE_ADDRNUMBER": SID,
                        "NAME1": "ALPHA LLC", "CITY1": "DENVER"}],
    "LFB1 - Company Code (Supplier)": [
        {"SOURCE_ID": SID, "BUKRS": "0601", "ZTERM": "Z000", "ZWELS": "A"},
        {"SOURCE_ID": SID, "BUKRS": "0432", "ZTERM": "Z000", "ZWELS": "A"}],
}


@pytest.fixture()
def consolidated(tmp_path):
    path = prefill_vendor(make_template(tmp_path / "t.xlsx"))
    tpl = BPTemplate(path)
    pre = tpl.snapshot()
    plan = tpl.append_rows(ROWS)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    return {"out": out, "plan": plan, "pre": pre}


def _edit(out, sheet, row, col=None, tech=None, value=None, delete_row=False):
    wb = openpyxl.load_workbook(out)
    ws = wb[sheet]
    if tech is not None:
        hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
        col = hdr[tech]
    if delete_row:
        ws.delete_rows(row)
    else:
        ws.cell(row=row, column=col, value=value)
    wb.save(out)
    wb.close()


def _errors(report):
    return [e for p in report["passes"] for e in p["errors"]]


def test_green_path(consolidated):
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {SID: ROWS})
    assert r["status"] == "verified", _errors(r)
    assert [p["pass"] for p in r["passes"]] == ["cell_check", "integrity", "roundtrip"]


def test_catches_value_shifted_one_column(consolidated):
    # simulate a column-shift: NAME1 value slides into the neighbour column
    out = consolidated["out"]
    wb = openpyxl.load_workbook(out)
    ws = wb["LFA1 - Supplier General"]
    hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    row = [c["row"] for c in consolidated["plan"]
           if c["sheet"] == "LFA1 - Supplier General"][0]
    ws.cell(row=row, column=hdr["NAME1"] + 1,
            value=ws.cell(row=row, column=hdr["NAME1"]).value)
    ws.cell(row=row, column=hdr["NAME1"]).value = None
    wb.save(out)
    wb.close()
    r = verify_output(out, consolidated["plan"], consolidated["pre"], {SID: ROWS})
    assert r["status"] == "blocked"
    assert any("NAME1" in e for e in _errors(r))


def test_catches_stray_value_in_unplanned_column(consolidated):
    # value duplicated into a column the plan never touched — B (planned cells
    # only) and C (row ranges + _COMMENT) are both blind; D's unexpected-column
    # check must fire
    row = min(c["row"] for c in consolidated["plan"]
              if c["sheet"] == "LFA1 - Supplier General")
    _edit(consolidated["out"], "LFA1 - Supplier General", row,
          tech="NAME2", value="ALPHA LLC")
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {SID: ROWS})
    assert r["status"] == "blocked"
    assert any("NAME2" in e and "planned nothing" in e for e in _errors(r))


def test_catches_deleted_appended_row(consolidated):
    row = min(c["row"] for c in consolidated["plan"]
              if c["sheet"] == "LFB1 - Company Code (Supplier)")
    _edit(consolidated["out"], "LFB1 - Company Code (Supplier)", row,
          delete_row=True)
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {SID: ROWS})
    assert r["status"] == "blocked"


def test_catches_mutated_preexisting_row(consolidated):
    # row 3 of LFA1 is the pre-existing vendor
    _edit(consolidated["out"], "LFA1 - Supplier General", 3,
          tech="NAME1", value="TAMPERED GMBH")
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {SID: ROWS})
    assert r["status"] == "blocked"
    assert any("MODIFIED" in e for e in _errors(r))


def test_catches_comment_poison_on_new_row(consolidated):
    row = min(c["row"] for c in consolidated["plan"]
              if c["sheet"] == "LFA1 - Supplier General")
    _edit(consolidated["out"], "LFA1 - Supplier General", row,
          tech="_COMMENT", value="oops")
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {SID: ROWS})
    assert r["status"] == "blocked"
    assert any("_COMMENT" in e for e in _errors(r))


def test_catches_stray_row_below_planned_range(consolidated):
    last = max(c["row"] for c in consolidated["plan"]
               if c["sheet"] == "BUT000 - General")
    _edit(consolidated["out"], "BUT000 - General", last + 3,
          tech="NAME_ORG1", value="STOWAWAY INC")
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {SID: ROWS})
    assert r["status"] == "blocked"
    assert any("stray" in e or "last used row" in e for e in _errors(r))


def test_pass_d_catches_consistent_plan_and_write_corruption(tmp_path):
    """Plan built from corrupted rows (city/name swapped) writes 'correctly'
    per passes B and C — only the round-trip against an independent
    re-derivation can catch it."""
    path = make_template(tmp_path / "t.xlsx")
    corrupted = {
        "LFA1 - Supplier General": [
            {"SOURCE_ID": SID, "NAME1": "DENVER", "LAND1": "US",
             "ORT01": "ALPHA LLC"}],  # swapped!
    }
    tpl = BPTemplate(path)
    pre = tpl.snapshot()
    plan = tpl.append_rows(corrupted)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    fresh = {"LFA1 - Supplier General": [
        {"SOURCE_ID": SID, "NAME1": "ALPHA LLC", "LAND1": "US",
         "ORT01": "DENVER"}]}
    r = verify_output(out, plan, pre, {SID: fresh})
    by_pass = {p["pass"]: p["errors"] for p in r["passes"]}
    assert by_pass["cell_check"] == []       # B is blind to this
    assert by_pass["integrity"] == []        # C is blind to this
    assert by_pass["roundtrip"]              # D catches it
    assert r["status"] == "blocked"


def test_roundtrip_checks_broadcast_consistency(tmp_path):
    path = make_template(tmp_path / "t.xlsx")
    tpl = BPTemplate(path)
    pre = tpl.snapshot()
    plan = tpl.append_rows(ROWS)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    _edit(out, "LFB1 - Company Code (Supplier)", 4, tech="ZTERM", value="Z030")
    r = verify_output(out, plan, pre, {SID: ROWS})
    assert r["status"] == "blocked"
    assert any("ZTERM" in e for e in _errors(r))


def test_missing_source_id_blocks_roundtrip(consolidated):
    r = verify_output(consolidated["out"], consolidated["plan"],
                      consolidated["pre"], {"": ROWS})
    assert r["status"] == "blocked"


def test_sheet_without_sid_column_notes_instead_of_false_block(tmp_path):
    # a sheet whose header has _COMMENT but no SOURCE_ID column gets rows
    # without a SID; pass D must skip it with a note, not block forever
    from consolidation_helpers import TEMPLATE_SHEETS
    sheets = dict(TEMPLATE_SHEETS)
    sheets["BUT100 - Role"] = ["_COMMENT", "_ACTION_CODE", "PARTNER", "RLTYP"]
    path = make_template(tmp_path / "t.xlsx", )
    # rebuild with modified BUT100
    path.unlink()
    make_template(path, sheets)
    tpl = BPTemplate(path)
    pre = tpl.snapshot()
    rows = dict(ROWS)
    rows["BUT100 - Role"] = [{"RLTYP": "FLVN00"}]
    plan = tpl.append_rows(rows)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    r = verify_output(out, plan, pre, {SID: rows})
    assert r["status"] == "verified", _errors(r)
    rt = [p for p in r["passes"] if p["pass"] == "roundtrip"][0]
    assert any("BUT100" in n for n in rt.get("notes", []))


def test_tin_masked_in_verify_error_messages(tmp_path):
    path = make_template(tmp_path / "t.xlsx")
    rows = {"LFA1 - Supplier General": [
        {"SOURCE_ID": SID, "NAME1": "ALPHA LLC", "LAND1": "US",
         "STCD1": "000-04-2016"}]}
    tpl = BPTemplate(path)
    pre = tpl.snapshot()
    plan = tpl.append_rows(rows)
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    _edit(out, "LFA1 - Supplier General", 3, tech="STCD1", value="111-11-1111")
    r = verify_output(out, plan, pre, {SID: rows})
    assert r["status"] == "blocked"
    blob = str(r)
    assert "000-04-2016" not in blob and "111-11-1111" not in blob
