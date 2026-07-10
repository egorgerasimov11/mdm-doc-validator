"""tin_bulk — the /ui/tax bulk US TIN validator (port of us-tax-number-validator).
Buckets + placeholder policy (punctuation/zeros -> Dummy) + workbook round-trip
+ route smoke. Tables come from rules/predicates.py — shared with W9-040/041."""
from pathlib import Path

from openpyxl import Workbook, load_workbook

from mdmdoc.tin_bulk import DUMMY, NOTAX, SUSPECT, VALID, classify, validate_workbook


def _v(raw):
    return classify(raw)[0]


def test_classify_valid_shapes():
    assert classify("36-1234567") == (VALID, "EIN", "EIN")
    assert classify("320-54-0693") == (VALID, "SSN", "SSN")
    assert classify("912-70-1234") == (VALID, "ITIN", "ITIN")
    assert classify("912-93-1234") == (VALID, "ITIN", "ITIN")          # ATIN group
    v, t, r = classify("212554321")                                    # bare 9
    assert v == VALID and "EIN" in t and "no separator" in r


def test_classify_structural_suspicious():
    v, _, r = classify("07-1234567")
    assert v == SUSPECT and "never-assigned IRS prefix" in r
    assert _v("000-12-3456") == SUSPECT
    assert _v("123-00-4567") == SUSPECT
    assert _v("123-45-0000") == SUSPECT
    v, _, r = classify("912-89-1234")
    assert v == SUSPECT and "ITIN group" in r
    assert _v("12-345678") == SUSPECT        # 8 digits -> wrong length
    assert _v("12345678901") == SUSPECT      # 11 digits


def test_classify_prefix_makes_suspicious():
    v, _, r = classify("US36-1234567")
    assert v == SUSPECT and 'extra "US" prefix' in r
    v, _, r = classify("EIN 36-1234567")
    assert v == SUSPECT and 'extra "EIN" prefix' in r


def test_classify_not_valid():
    assert _v("DE137196337") == NOTAX        # foreign VAT
    assert _v("16R481") == NOTAX             # letters
    assert _v("12345") == NOTAX              # too short
    assert _v("123456789012") == NOTAX       # >= 12 digits
    assert _v("") == NOTAX


def test_classify_dummy_policy():
    assert _v(".") == DUMMY                  # punctuation placeholder (per Egor)
    assert _v("0") == DUMMY
    assert _v("000000000") == DUMMY
    assert _v("999999999") == DUMMY
    assert _v("XXXXXXX") == DUMMY
    assert _v("US99999999999") == DUMMY      # long 0/9 mix with US prefix
    assert _v("123456789") == DUMMY          # known fake
    assert _v("078-05-1120") == DUMMY        # SSA advertising SSN


def _mini_xlsx(path: Path) -> None:
    wb = Workbook()
    sh = wb.active
    sh.append(["Business Partner", "Tax Number Category", "Tax number"])
    for bp, cat, tin in (("100", "US2", "36-1234567"),   # valid EIN
                         ("101", "US1", "320-54-0693"),  # valid SSN
                         ("102", "US0", "07-1234567"),   # dead prefix
                         ("103", "US4", "XXXXXXX"),      # masked
                         ("104", "US3", "."),            # punctuation
                         ("105", "US0", "DE137196337")): # foreign
        sh.append([bp, cat, tin])
    wb.save(path)


def test_validate_workbook_roundtrip(tmp_path):
    src = tmp_path / "export.xlsx"
    out = tmp_path / "out.xlsx"
    _mini_xlsx(src)
    summary = validate_workbook(src, out)
    assert summary["total"] == 6
    assert summary["counts"] == {VALID: 2, SUSPECT: 1, NOTAX: 1, DUMMY: 2}
    assert summary["columns"]["tax"] == "Tax number"
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Valid", "Suspicious", "Not valid", "Dummy-Masked"}
    valid_rows = list(wb["Valid"].iter_rows(min_row=2, values_only=True))
    assert {r[0] for r in valid_rows} == {"100", "101"}
    assert all(len(r) == 6 for r in valid_rows)


def test_validate_workbook_explicit_columns(tmp_path):
    src = tmp_path / "odd.xlsx"
    wb = Workbook()
    sh = wb.active
    sh.append(["Nr", "Wert"])
    sh.append(["7", "36-1234567"])
    wb.save(src)
    out = tmp_path / "out.xlsx"
    s = validate_workbook(src, out, col="B", id_col="A")
    assert s["total"] == 1 and s["counts"][VALID] == 1


def test_tax_routes_smoke(tmp_path, monkeypatch):
    from fastapi.testclient import TestClient

    from mdmdoc.server import ui
    from mdmdoc.server.app import create_app
    monkeypatch.setattr(ui, "_TAX_DIR", tmp_path / "tax")
    client = TestClient(create_app("full"))
    assert client.get("/ui/tax").status_code == 200

    src = tmp_path / "mini.xlsx"
    _mini_xlsx(src)
    r = client.post("/ui/tax/validate",
                    files={"file": ("mini.xlsx", src.read_bytes(),
                                    "application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet")})
    assert r.status_code == 200 and "Result — 6 rows" in r.text
    name = next(p.name for p in (tmp_path / "tax").glob("*_VALIDATED.xlsx"))
    dl = client.get(f"/ui/tax/download/{name}")
    assert dl.status_code == 200 and dl.content[:2] == b"PK"
    assert client.get("/ui/tax/download/../secrets.xlsx").status_code in (404, 422)
    assert client.get("/ui/tax/download/nope.xlsx").status_code == 404
