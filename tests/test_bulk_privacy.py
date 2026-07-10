"""V-wave: privacy posture — runs/ bulk artifacts never carry a full tax
number; the full-value workbook lives ONLY in inbox/."""
import json

from openpyxl import Workbook

from mdmdoc import config
from mdmdoc.bulk import run_bulk

FULL_TIN = "12-3456789"
FULL_DE = "DE137196337"


def _input(tmp_path):
    p = tmp_path / "in.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Business Partner", "Tax Number Category", "Tax Number",
               "Tax Number Long", "Country"])
    ws.append(["1", "US2", FULL_TIN, "", "US"])
    ws.append(["2", "US0", FULL_DE, "", "US"])
    ws.append(["3", "US1", "999-99-0000", "", "US"])   # structurally invalid
    wb.save(p)
    return p


def test_runs_artifacts_masked_inbox_full(tmp_path, monkeypatch):
    from mdmdoc import runstore
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)

    res, arts = run_bulk(_input(tmp_path))
    bid = arts["bulk_id"]

    blob = (tmp_path / "runs" / bid / "bulk_report.json").read_text()
    blob += (tmp_path / "runs" / bid / "bulk_report.md").read_text()
    # full tax numbers must NOT appear in leak-gated artifacts...
    assert FULL_TIN not in blob
    assert FULL_TIN.replace("-", "") not in blob
    assert "999-99-0000" not in blob
    # ...while the wrong-country reason still identifies the problem (masked)
    rep = json.loads((tmp_path / "runs" / bid / "bulk_report.json").read_text())
    t03 = [r for r in rep["problem_rows"] if "BULK-T03" in r["rules"]]
    assert t03 and "German VAT" in t03[0]["reasons"][0]

    # the operator's workbook in inbox/ DOES carry the full values
    from openpyxl import load_workbook
    wb = load_workbook(arts["result_xlsx"])
    data = "\n".join(str(c.value) for row in wb["Data"].iter_rows()
                     for c in row if c.value is not None)
    assert FULL_TIN in data and FULL_DE in data
    # and it lives under inbox/, never under runs/
    assert str(tmp_path / "inbox") in arts["result_xlsx"]
