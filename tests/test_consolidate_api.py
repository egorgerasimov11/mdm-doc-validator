"""Consolidation tab routes — extract, template, consolidate, dedup, download."""
from __future__ import annotations

import io

from consolidation_helpers import (
    make_americas_form,
    make_template,
    needs_converter,
    prefill_vendor,
)
from mdmdoc import config

pytestmark = needs_converter


def _client(monkeypatch, tmp_path):
    from fastapi.testclient import TestClient

    from mdmdoc import runstore
    from mdmdoc.server.app import create_app
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)
    return TestClient(create_app("full"))


def _bytes_of(path):
    return io.BytesIO(path.read_bytes())


def _new_case(client):
    r = client.post("/ui/consolidation/new", follow_redirects=False)
    assert r.status_code == 303
    return r.headers["location"].rstrip("/").rsplit("/", 1)[-1]


def _upload(client, case, route, path, name=None, data=None):
    return client.post(
        f"/ui/consolidation/{case}/{route}",
        files={"file": (name or path.name, _bytes_of(path),
                        "application/octet-stream")},
        data=data or {},
        follow_redirects=False)


def test_full_flow_extract_template_consolidate_download(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    form = make_americas_form(tmp_path / "form.xlsm")
    tpl = make_template(tmp_path / "BusinessPartnerTemplate.xlsx")
    case = _new_case(client)

    assert _upload(client, case, "extract", form).status_code == 303
    page = client.get(f"/ui/consolidation/{case}").text
    assert "SETH FAKESON" in page and "NEW_" in page and "uploaded" in page

    assert _upload(client, case, "template", tpl).status_code == 303

    # warnings (8-digit ABA + SORTL) block without confirm — nothing written
    r = client.post(f"/ui/consolidation/{case}/consolidate")
    assert "nothing was written" in r.text
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 409

    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    page = client.get(f"/ui/consolidation/{case}").text
    assert "verified" in page and "consolidated_v001.xlsx" in page

    dl = client.get(f"/ui/consolidation/{case}/download")
    assert dl.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["LFA1 - Supplier General"]
    hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    assert ws.cell(row=3, column=hdr["NAME1"]).value == "SETH FAKESON"
    wb.close()


def test_append_twice_output_feeds_next_consolidate(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    tpl = make_template(tmp_path / "tpl.xlsx")
    case = _new_case(client)
    _upload(client, case, "template", tpl)

    form1 = make_americas_form(tmp_path / "f1.xlsm")
    form2 = make_americas_form(tmp_path / "f2.xlsm",
                               vendor_name="OTHER FAKESON")
    _upload(client, case, "extract", form1, name="f1.xlsm")
    client.post(f"/ui/consolidation/{case}/consolidate",
                data={"confirm_warnings": "on"})

    _upload(client, case, "extract", form2, name="f2.xlsm")
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303

    dl = client.get(f"/ui/consolidation/{case}/download")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["LFA1 - Supplier General"]
    hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    names = {ws.cell(row=r_, column=hdr["NAME1"]).value for r_ in (3, 4)}
    sids = {ws.cell(row=r_, column=hdr["SOURCE_ID"]).value for r_ in (3, 4)}
    wb.close()
    assert names == {"SETH FAKESON", "OTHER FAKESON"}
    assert len(sids) == 2  # unique SOURCE_IDs


def test_duplicate_vendor_blocked_then_override(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    tpl = prefill_vendor(make_template(tmp_path / "tpl.xlsx"),
                         source_id="OLD_1", name="SETH FAKESON")
    form = make_americas_form(tmp_path / "form.xlsm")
    case = _new_case(client)
    _upload(client, case, "extract", form)
    _upload(client, case, "template", tpl)

    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"})
    assert "duplicate vendor" in r.text
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 409

    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on", "override_dup": "on"},
                    follow_redirects=False)
    assert r.status_code == 303
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 200


def test_reused_source_id_hard_blocked(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    tpl = prefill_vendor(make_template(tmp_path / "tpl.xlsx"),
                         source_id="38300", name="SOMEBODY ELSE")
    # form with Vendor Number 38300 -> SOURCE_ID collides with the template
    form = make_americas_form(tmp_path / "form.xlsm", vendor_number="38300")
    case = _new_case(client)
    _upload(client, case, "extract", form)
    _upload(client, case, "template", tpl)
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on", "override_dup": "on"})
    assert "already exists" in r.text
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 409


def test_not_a_form_rejected(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    case = _new_case(client)
    not_form = make_template(tmp_path / "actually_a_template.xlsx")
    r = _upload(client, case, "extract", not_form)
    assert "does not look like a vendor request form" in r.text


def test_foreign_template_rejected(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    case = _new_case(client)
    import openpyxl
    p = tmp_path / "foreign.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "nope"
    wb.save(p)
    wb.close()
    r = _upload(client, case, "template", p)
    assert "not a" in r.text and "template" in r.text


def test_source_id_edit_and_validation(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    form = make_americas_form(tmp_path / "form.xlsm")
    case = _new_case(client)
    _upload(client, case, "extract", form)
    r = client.post(f"/ui/consolidation/{case}/vendor/v01/source-id",
                    data={"source_id": "bad id!"})
    assert "SOURCE_ID" in r.text
    r = client.post(f"/ui/consolidation/{case}/vendor/v01/source-id",
                    data={"source_id": "VEND_A1"}, follow_redirects=False)
    assert r.status_code == 303
    assert "VEND_A1" in client.get(f"/ui/consolidation/{case}").text


def test_unknown_case_404(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    assert client.get("/ui/consolidation/../etc").status_code in (404, 400)
    assert client.get("/ui/consolidation/c00000000_000000_dead").status_code == 404


def test_form_extract_json_never_holds_full_tin(monkeypatch, tmp_path):
    # review finding: the raw extract half of extract_vNN.json carried the
    # full TIN; masked_extract must strip it under every policy
    client = _client(monkeypatch, tmp_path)
    form = make_americas_form(tmp_path / "form.xlsm")  # SSN 000-04-2016
    case = _new_case(client)
    _upload(client, case, "extract", form)
    case_dir = tmp_path / "runs" / "consolidation" / case
    hits = [p.name for p in case_dir.rglob("*.json")
            if "000-04-2016" in p.read_text(encoding="utf-8")]
    assert hits == [], hits


def test_same_batch_duplicate_vendor_blocked(monkeypatch, tmp_path):
    # two new-vendor forms with the same name in ONE consolidate call must
    # trip the dup gate even though their SOURCE_IDs differ
    client = _client(monkeypatch, tmp_path)
    tpl = make_template(tmp_path / "tpl.xlsx")
    case = _new_case(client)
    _upload(client, case, "template", tpl)
    f1 = make_americas_form(tmp_path / "f1.xlsm")
    f2 = make_americas_form(tmp_path / "f2.xlsm")
    _upload(client, case, "extract", f1, name="f1.xlsm")
    _upload(client, case, "extract", f2, name="f2.xlsm")
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"})
    assert "duplicate vendor" in r.text
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 409


def test_template_replace_unconsolidates_vendors(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    tpl = make_template(tmp_path / "tpl.xlsx")
    case = _new_case(client)
    form = make_americas_form(tmp_path / "form.xlsm")
    _upload(client, case, "extract", form)
    _upload(client, case, "template", tpl)
    client.post(f"/ui/consolidation/{case}/consolidate",
                data={"confirm_warnings": "on"})
    assert "consolidated_v001.xlsx" in client.get(f"/ui/consolidation/{case}").text
    # replace the template — the vendor's rows lived in the discarded output
    tpl2 = make_template(tmp_path / "tpl2.xlsx")
    r = _upload(client, case, "template", tpl2,
                data={"confirm_replace": "on"})
    assert r.status_code == 303
    page = client.get(f"/ui/consolidation/{case}").text
    assert "Consolidate 1 vendor(s)" in page  # back to pending
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 409


def test_source_id_locked_after_consolidation(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    tpl = make_template(tmp_path / "tpl.xlsx")
    case = _new_case(client)
    form = make_americas_form(tmp_path / "form.xlsm")
    _upload(client, case, "extract", form)
    _upload(client, case, "template", tpl)
    client.post(f"/ui/consolidation/{case}/consolidate",
                data={"confirm_warnings": "on"})
    r = client.post(f"/ui/consolidation/{case}/vendor/v01/source-id",
                    data={"source_id": "CHANGED_1"})
    assert "cannot change" in r.text
