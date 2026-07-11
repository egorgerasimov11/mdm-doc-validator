"""Hide non-key EMPTY columns in the output so SAP validates only populated
and key columns. SAP validates every VISIBLE column: an empty timestamp
(UPTIM) / GUID (BP_BANK_GUID) is a hard error and an empty date a warning, so
empty non-key columns are hidden (SAP ignores hidden columns on import). KEY
columns (underlined) stay visible even when blank."""
from __future__ import annotations

import zipfile

import openpyxl
from openpyxl.utils import get_column_letter

from consolidation_helpers import make_template
from mdmdoc.consolidation.template_io import BPTemplate

ROWS = {
    "LFA1 - Supplier General": [{"SOURCE_ID": "N1", "NAME1": "上海外服 LLC",
                                 "LAND1": "CN"}],
}


def _letter(ws, tech):
    for c in ws[2]:
        if c.value and str(c.value).strip() == tech:
            return get_column_letter(c.column)
    raise KeyError(tech)


def _hidden(ws, tech):
    return bool(ws.column_dimensions[_letter(ws, tech)].hidden)


def _out(tmp_path, rows=ROWS, **save_kw):
    tpl = BPTemplate(make_template(tmp_path / "t.xlsx"))
    tpl.append_rows(rows)
    out = tpl.save_to(tmp_path / "out.xlsx", **save_kw)
    tpl.close()
    return out


def test_non_key_empty_columns_hidden(tmp_path):
    ws = openpyxl.load_workbook(_out(tmp_path))["LFA1 - Supplier General"]
    assert _hidden(ws, "TELF1")           # empty non-key → hidden
    assert _hidden(ws, "STCD1")           # empty non-key → hidden


def test_populated_columns_stay_visible(tmp_path):
    ws = openpyxl.load_workbook(_out(tmp_path))["LFA1 - Supplier General"]
    for tech in ("SOURCE_ID", "NAME1", "LAND1"):
        assert not _hidden(ws, tech), tech


def test_key_column_stays_visible_when_empty(tmp_path):
    # ADRC has an underlined NATION key; leave it blank on the appended row
    from test_consolidate_required_fields import _template_with_address_keys
    tpl = BPTemplate(_template_with_address_keys(tmp_path))
    tpl.append_rows({"ADRC - Address": [
        {"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "N1", "NAME1": "上海外服 LLC"}]})
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    ws = openpyxl.load_workbook(out)["ADRC - Address"]
    assert not _hidden(ws, "NATION")      # empty KEY column NOT hidden
    assert not _hidden(ws, "NAME1")       # populated → visible
    assert not _hidden(ws, "CITY1")       # empty but WHITELISTED business field → visible
    assert _hidden(ws, "FAX_NUMBER")      # empty non-key, non-whitelisted → hidden


def test_hide_empty_false_keeps_columns_visible(tmp_path):
    ws = openpyxl.load_workbook(
        _out(tmp_path, hide_empty=False))["LFA1 - Supplier General"]
    assert not _hidden(ws, "TELF1")       # opt-out (base_cleared path)


def test_hidden_output_keeps_values_and_shared_strings(tmp_path):
    out = _out(tmp_path)
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["LFA1 - Supplier General"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    got = [ws.cell(row=r, column=h["NAME1"]).value
           for r in range(3, ws.max_row + 1)]
    assert "上海外服 LLC" in got
    wb.close()
    z = zipfile.ZipFile(out)
    inline = sum(z.read(n).count(b"inlineStr")
                 for n in z.namelist() if n.startswith("xl/worksheets/"))
    assert inline == 0                    # hiding didn't reintroduce inline
