"""Exact-cell mapping assertions on a synthesized SETH_DEVRIES-shaped form.

Pins every transform the real Americas packet exercises: fan-out 0601/0432
and U001/U002, code stripping (Z000/A/01/ZKTV/B344001000/Z01/VTP_DOM_BI),
USA->US, MI passthrough, ENG->EN, SSN dashes kept, DFKKBPTAXNUM US1 rows,
HCE search term, SOURCE_ID stamped everywhere, and the two expected pass-A
warnings (SORTL length + 8-digit ABA with checksum-valid zero-pad).
"""
from __future__ import annotations

import openpyxl
import pytest

from consolidation_helpers import (
    cell,
    make_americas_form,
    make_template,
    needs_converter,
)
from mdmdoc.consolidation import convert, plan as planmod
from mdmdoc.consolidation.template_io import BPTemplate

pytestmark = needs_converter

SID = "NEW_20260710_01"


@pytest.fixture()
def written(tmp_path):
    form = make_americas_form(tmp_path / "form.xlsm")
    tpl = BPTemplate(make_template(tmp_path / "tpl.xlsx"))
    tpl.validate()
    built = convert.build_vendor_rows(form, tpl, source_id=SID)
    assert built["errors"] == []
    cells = tpl.plan_rows(built["rows"])
    review = planmod.review(cells, kind="form", source_id=SID)
    tpl.append_rows(built["rows"])
    out = tpl.save_to(tmp_path / "out.xlsx")
    tpl.close()
    wb = openpyxl.load_workbook(out)
    yield {"wb": wb, "built": built, "review": review, "form": form}
    wb.close()


def test_lfa1_row3(written):
    wb = written["wb"]
    s = "LFA1 - Supplier General"
    assert cell(wb, s, 3, "NAME1") == "SETH FAKESON"
    assert cell(wb, s, 3, "LAND1") == "US"
    assert cell(wb, s, 3, "REGIO") == "MI"
    assert cell(wb, s, 3, "PSTLZ") == "49506"
    assert cell(wb, s, 3, "ORT01") == "GRAND RAPIDS"
    assert cell(wb, s, 3, "STRAS") == "1602 FAKE RD SE"
    assert cell(wb, s, 3, "KTOKK") == "ZKTV"
    assert cell(wb, s, 3, "STCD1") == "000-04-2016"   # dashes kept, as entered
    assert cell(wb, s, 3, "SPRAS") == "EN"            # ENG -> EN
    assert cell(wb, s, 3, "SOURCE_ID") == SID
    assert cell(wb, s, 3, "LIFNR") is None            # new vendor: no number
    assert cell(wb, s, 4, "NAME1") is None            # exactly one row


def test_but000_row3(written):
    wb = written["wb"]
    s = "BUT000 - General"
    assert cell(wb, s, 3, "NAME_ORG1") == "SETH FAKESON"
    assert cell(wb, s, 3, "BU_GROUP") == "ZKTV"
    assert cell(wb, s, 3, "BU_SORT1") == "SETH FAKESON"
    assert cell(wb, s, 3, "BU_SORT2") == "HCE"
    assert cell(wb, s, 3, "BU_LANGU") == "EN"
    assert cell(wb, s, 3, "SOURCE_ID") == SID
    assert cell(wb, s, 3, "PARTNER") is None


def test_lfb1_fanout_two_rows_identical_broadcasts(written):
    wb = written["wb"]
    s = "LFB1 - Company Code (Supplier)"
    assert {cell(wb, s, 3, "BUKRS"), cell(wb, s, 4, "BUKRS")} == {"0601", "0432"}
    for r in (3, 4):
        assert cell(wb, s, r, "ZTERM") == "Z000"
        assert cell(wb, s, r, "ZWELS") == "A"
        assert cell(wb, s, r, "QSREC") == "01"      # "1 - Individual" -> 01
        assert cell(wb, s, r, "AKONT") == "B344001000"
        assert cell(wb, s, r, "ZUAWA") == "Z01"
        assert cell(wb, s, r, "FDGRV") == "VTP_DOM_BI"
        assert cell(wb, s, r, "SOURCE_ID") == SID
        assert cell(wb, s, r, "_ACTION_CODE") == "I"
    assert cell(wb, s, 5, "BUKRS") is None


def test_lfm1_fanout(written):
    wb = written["wb"]
    s = "LFM1 - Purchasing Org Data"
    assert {cell(wb, s, 3, "EKORG"), cell(wb, s, 4, "EKORG")} == {"U001", "U002"}
    for r in (3, 4):
        assert cell(wb, s, r, "WAERS") == "USD"
        assert cell(wb, s, r, "ZTERM") == "Z000"


def test_but0bk_row3(written):
    wb = written["wb"]
    s = "BUT0BK - Bank Account"
    assert cell(wb, s, 3, "BKVID") == "0001"
    assert cell(wb, s, 3, "BANKS") == "US"
    assert cell(wb, s, 3, "BANKL") == "71000013"   # as entered; pass A warns
    assert cell(wb, s, 3, "BANKN") == "000661570"  # leading zeros preserved


def test_dfkkbptaxnum_us1(written):
    wb = written["wb"]
    s = "DFKKBPTAXNUM - Tax Number"
    assert cell(wb, s, 3, "TAXTYPE") == "US1"
    assert cell(wb, s, 3, "TAXNUM") == "000-04-2016"
    assert cell(wb, s, 3, "SOURCE_ID") == SID


def test_but100_static_role(written):
    wb = written["wb"]
    assert cell(wb, "BUT100 - Role", 3, "RLTYP") == "FLVN00"


def test_adr6_email(written):
    wb = written["wb"]
    assert cell(wb, "ADR6 - E-Mail", 3, "SMTP_ADDR") == "seth.fakeson@example.org"


def test_pass_a_expected_warnings_only(written):
    msgs = [w["message"] for w in written["review"]["warnings"]]
    assert any("SORTL" in m for m in msgs)                 # 12-char search term
    assert any("071000013" in m for m in msgs)             # ABA zero-pad hint
    assert written["review"]["errors"] == []


def test_placeholders_never_reach_workbook(written):
    wb = written["wb"]
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=3, values_only=True):
            for v in row:
                if v is None:
                    continue
                s = str(v)
                assert "(Required" not in s and s not in ("N/A", "Ñ"), \
                    f"placeholder leaked: {s!r} in {ws.title}"


def test_comment_empty_everywhere(written):
    wb = written["wb"]
    for ws in wb.worksheets:
        hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
        col = hdr.get("_COMMENT")
        if not col:
            continue
        for r in range(3, ws.max_row + 1):
            assert not ws.cell(row=r, column=col).value


def test_source_id_on_every_written_row(written):
    built = written["built"]
    for sheet, rows in built["rows"].items():
        for row in rows:
            assert row.get("SOURCE_ID") == SID or row.get("SOURCE_ADDRNUMBER") == SID, \
                (sheet, row)


def test_coverage_no_field_lost(written):
    extract = convert.extract_form(written["form"])
    cov = convert.coverage(extract["fields"], written["built"]["rows"],
                           written["built"]["unmapped"])
    lost = [c for c in cov if c["status"] == "not_loaded"]
    assert lost == [], lost
