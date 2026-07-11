"""Task 1 — export of a validated document run into a consolidation case.

Gate (Egor's decision): ACCEPT freely, WARNING behind an explicit confirm,
REVIEW/REJECT never. Full values only from full-policy artifacts; the leak
gate proves the full TIN lands in the OUTPUT WORKBOOK only, never in case
JSON artifacts.
"""
from __future__ import annotations

import io
import json

from consolidation_helpers import make_template, needs_converter
from mdmdoc import config

pytestmark = needs_converter

FAKE_TIN = "000-04-2016"
FAKE_ACCT = "000661570"


def _client(monkeypatch, tmp_path):
    from fastapi.testclient import TestClient

    from mdmdoc import runstore
    from mdmdoc.server.app import create_app
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)
    monkeypatch.setenv("MDMDOC_BANK_VALUES", "full")
    monkeypatch.setenv("MDMDOC_TIN_VALUES", "full")
    return TestClient(create_app("full"))


def _make_run(tmp_path, run_id, doc_class, verdict, fields, tpl_rows=None):
    d = tmp_path / "runs" / run_id
    d.mkdir(parents=True, exist_ok=True)
    (d / "meta.json").write_text(json.dumps(
        {"run_id": run_id, "file_name": f"{run_id}.pdf", "doc_class": doc_class}))
    (d / "report.json").write_text(json.dumps(
        {"schema": "mdmdoc.v1", "doc_class": doc_class, "verdict": verdict}))
    (d / "extraction.json").write_text(json.dumps(
        {"doc_class": doc_class, "fields": fields, "provenance": {}}))
    if tpl_rows is not None:
        (d / "template_compare.json").write_text(json.dumps(tpl_rows))


W9_FIELDS = {
    "line1_name": "SETH FAKESON",
    "tin_type": "SSN",
    "tin_raw": {"masked": "XXX-XX-2016", "present": True, "value": FAKE_TIN,
                "length": 11},
    "address_street": "1602 FAKE RD SE",
    "address_city_state_zip": "GRAND RAPIDS, MI 49506",
}

BANK_FIELDS = {
    "account_holder": "SETH FAKESON",
    "bank_country": "US",
    "routing_aba": {"masked": "0710*****", "present": True, "value": "071000013",
                    "length": 9},
    "account_number": {"masked": "***570", "present": True, "value": FAKE_ACCT,
                       "length": 9},
    "bank_name": "FAKE BANK",
}


def test_accept_w9_exports_fragment(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    _make_run(tmp_path, "runw9ok", "w9", "ACCEPT", W9_FIELDS)
    r = client.post("/ui/consolidation/from-run/runw9ok", follow_redirects=False)
    assert r.status_code == 303
    case = r.headers["location"].rstrip("/").rsplit("/", 1)[-1]
    page = client.get(f"/ui/consolidation/{case}").text
    assert "doc" in page and "extracted" in page

    # consolidate the fragment into a template — full TIN lands in the file
    tpl = make_template(tmp_path / "tpl.xlsx")
    client.post(f"/ui/consolidation/{case}/template",
                files={"file": ("tpl.xlsx", io.BytesIO(tpl.read_bytes()),
                                "application/octet-stream")})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    follow_redirects=False)
    if r.status_code != 303:  # warnings need confirm
        r = client.post(f"/ui/consolidation/{case}/consolidate",
                        data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    dl = client.get(f"/ui/consolidation/{case}/download")
    assert dl.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["LFA1 - Supplier General"]
    hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    assert ws.cell(row=3, column=hdr["STCD1"]).value == FAKE_TIN
    tax = wb["DFKKBPTAXNUM - Tax Number"]
    hdr2 = {str(c.value).strip(): c.column for c in tax[2] if c.value}
    assert tax.cell(row=3, column=hdr2["TAXTYPE"]).value == "US1"
    wb.close()

    # leak gate: full TIN exists ONLY in the workbook, not in case JSONs
    case_dir = tmp_path / "runs" / "consolidation" / case
    for p in case_dir.rglob("*.json"):
        assert FAKE_TIN not in p.read_text(encoding="utf-8"), p.name


def test_reject_blocked(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    _make_run(tmp_path, "runrej", "bank", "REJECT", BANK_FIELDS)
    r = client.post("/ui/consolidation/from-run/runrej")
    assert r.status_code == 409
    assert "REJECT" in r.text


def test_review_blocked(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    _make_run(tmp_path, "runrev", "bank", "NEED_MANUAL_REVIEW", BANK_FIELDS)
    assert client.post("/ui/consolidation/from-run/runrev").status_code == 409


def test_warning_needs_confirm(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    _make_run(tmp_path, "runwarn", "bank", "WARNING", BANK_FIELDS)
    assert client.post("/ui/consolidation/from-run/runwarn").status_code == 409
    r = client.post("/ui/consolidation/from-run/runwarn",
                    data={"confirm": "on"}, follow_redirects=False)
    assert r.status_code == 303


def test_masked_artifact_blocked_with_advice(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    masked = dict(W9_FIELDS)
    masked["tin_raw"] = {"masked": "XXX-XX-2016", "present": True, "length": 11}
    _make_run(tmp_path, "runmask", "w9", "ACCEPT", masked)
    r = client.post("/ui/consolidation/from-run/runmask", follow_redirects=False)
    assert r.status_code == 303  # case created; vendor carries the error
    case = r.headers["location"].rstrip("/").rsplit("/", 1)[-1]
    page = client.get(f"/ui/consolidation/{case}").text
    assert "Re-analyze" in page


def test_doc_form_mismatch_blocks_export(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    _make_run(tmp_path, "runmm", "bank", "ACCEPT", BANK_FIELDS, tpl_rows=[
        {"field": "Bank Account", "doc": "***570", "sap": "***571",
         "status": "MISMATCH", "note": ""}])
    r = client.post("/ui/consolidation/from-run/runmm")
    assert r.status_code == 409
    assert "MISMATCH" in r.text


def test_blank_tin_type_infers_ein_from_shape(monkeypatch, tmp_path):
    # review finding: blank tin_type used to file an EIN-shaped TIN under
    # STCD1/US1; shape inference must route XX-XXXXXXX to STCD2/US2
    from mdmdoc.consolidation import docrows

    _client(monkeypatch, tmp_path)  # sets env + RUNS_DIR
    fields = dict(W9_FIELDS)
    fields["tin_type"] = ""
    fields["tin_raw"] = {"masked": "XX-XXX6789", "present": True,
                         "value": "12-3456789", "length": 10}
    _make_run(tmp_path, "runein", "w9", "ACCEPT", fields)

    class _Cols:
        def has_sheet(self, s):
            return True

        def column_for(self, s, t):
            return 1

    built = docrows.rows_from_run("runein", _Cols(), source_id="N1")
    lfa1 = built["rows"]["LFA1 - Supplier General"][0]
    assert lfa1.get("STCD2") == "12-3456789"
    assert "STCD1" not in lfa1
    tax = built["rows"]["DFKKBPTAXNUM - Tax Number"][0]
    assert tax["TAXTYPE"] == "US2"


def test_masked_policy_deployment_blocked(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    monkeypatch.setenv("MDMDOC_BANK_VALUES", "masked")
    _make_run(tmp_path, "runpol", "bank", "ACCEPT", BANK_FIELDS)
    r = client.post("/ui/consolidation/from-run/runpol")
    assert r.status_code == 409
    assert "masked" in r.text
