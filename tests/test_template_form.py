"""D8: template compare — the request-form workbook parses generically
(labels anywhere, value to the right), feeds the SAME char-by-char comparer
as SAP with TPL-namespaced findings, fails closed, keeps secrets secret."""
import fitz
import pytest

from mdmdoc import config, template_form
from mdmdoc.pipeline import run_check


def _workbook(path, rows_by_sheet: dict):
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for sheet, rows in rows_by_sheet.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = sheet
        first = False
        for r, row in enumerate(rows, start=1):
            for c, v in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


AMERICAS = {
    "1. General Info": [
        ["", "Vendor Name:", "SETH FAKESON"],
    ],
    "2. Vendor Details": [
        # labels in column D, values in E — replicating the real layout
        ["", "", "", "Bank Name", "CHASE BANK"],
        ["", "", "", "Bank Country", "USA"],
        ["", "", "", "Bank Key (ABA Routing Code)", "072000326"],
        ["", "", "", "Bank Account Number", "683661570"],
        ["", "", "", "IBAN Number", None],
        ["", "", "", "BIC/Swift Code", None],
        ["", "", "", "Tax Number 1", "000-04-2016"],
        ["", "", "", "Payment Method", "A - ACH"],
        ["", "", "", "Order Currency", "USD"],
    ],
}


def test_parse_vertical_form(tmp_path):
    wb = _workbook(tmp_path / "americas.xlsm", AMERICAS)
    fields, prov, hits = template_form.parse(wb)
    assert hits >= 7
    assert fields["bank_name"] == "CHASE BANK"
    assert fields["bank_key"] == "072000326"
    assert fields["bank_account"] == "683661570"
    assert fields["vendor_name"] == "SETH FAKESON"
    assert fields["iban"] == ""                       # label found, value empty
    assert prov["bank_key"].startswith("2. Vendor Details!")
    assert template_form.looks_like_request_form(wb) is True


def test_alias_tolerance(tmp_path):
    wb = _workbook(tmp_path / "eu.xlsx", {"Form": [
        ["Beneficiary Name", "Fake Corp GmbH"],
        ["IBAN", "DE89 3704 0044 0532 0130 00"],
        ["SWIFT Code", "COBADEFFXXX"],
    ]})
    fields, _, hits = template_form.parse(wb)
    assert hits == 3
    assert fields["account_holder"] == "Fake Corp GmbH"
    assert fields["swift_bic"] == "COBADEFFXXX"


def test_to_sap_fields_holder_falls_back_to_vendor_name():
    out = template_form.to_sap_fields({"vendor_name": "SETH FAKESON",
                                       "bank_key": "072000326"})
    assert out["account_holder"] == "SETH FAKESON"
    assert out["bank_key"] == "072000326"


def _pdf(tmp_path, text, name="doc.pdf"):
    p = tmp_path / name
    d = fitz.open()
    pg = d.new_page()
    y = 80
    for line in text.splitlines():
        pg.insert_text((72, y), line, fontsize=10)
        y += 16
    d.save(p)
    d.close()
    return p


def test_template_compare_mismatch_folds_to_nmr(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    wb = _workbook(tmp_path / "form.xlsx", {"Form": [
        ["Bank Name", "Commerzbank AG"],
        ["IBAN Number", "DE44 5001 0517 5407 3249 31"],   # DIFFERENT from the doc
        ["Beneficiary Name", "Fake Corp GmbH"],
    ]})
    doc = _pdf(tmp_path, "Bank confirmation letter\n"
                         "This letter is to confirm the account details below.\n"
                         "Account holder: Fake Corp GmbH\n"
                         "IBAN: DE89 3704 0044 0532 0130 00\n"
                         "Bank: Commerzbank AG")
    res = run_check(doc, "bank", use_vision=False, engine="deterministic",
                    enforce_approvals=False, template_path=wb)
    ids = {f.rule_id for f in res.findings}
    assert "TPL-001" in ids                            # IBAN mismatch, TPL namespace
    assert res.verdict == "NEED_MANUAL_REVIEW"
    rows = res.pub.get("template_compare") or []
    assert any(r["status"] == "MISMATCH" for r in rows)
    md = (config.RUNS_DIR / res.run_id / "reasoning.md").read_text()
    assert "Template (request form) comparison" in md
    tpl_art = (config.RUNS_DIR / res.run_id / "template_compare.json")
    assert tpl_art.exists()


def test_template_match_confirms(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    wb = _workbook(tmp_path / "form.xlsx", {"Form": [
        ["Bank Name", "Commerzbank AG"],
        ["IBAN Number", "DE89 3704 0044 0532 0130 00"],
        ["Beneficiary Name", "Fake Corp GmbH"],
    ]})
    doc = _pdf(tmp_path, "Bank confirmation letter\n"
                         "This letter is to confirm the account details below.\n"
                         "Account holder: Fake Corp GmbH\n"
                         "IBAN: DE89 3704 0044 0532 0130 00\n"
                         "Bank: Commerzbank AG")
    res = run_check(doc, "bank", use_vision=False, engine="deterministic",
                    enforce_approvals=False, template_path=wb)
    assert not any(f.rule_id.startswith("TPL-0") and f.severity == "CRITICAL"
                   for f in res.findings)


def test_not_a_form_skips_with_tpl015(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    wb = _workbook(tmp_path / "junk.xlsx", {"Sheet": [["random", "data"]]})
    doc = _pdf(tmp_path, "Bank confirmation letter\nIBAN DE89 3704 0044 0532 0130 00")
    res = run_check(doc, "bank", use_vision=False, engine="deterministic",
                    enforce_approvals=False, template_path=wb)
    assert any(f.rule_id == "TPL-015" for f in res.findings)


def test_fail_closed_tpl014(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(template_form, "parse",
                        lambda p: (_ for _ in ()).throw(RuntimeError("boom")))
    wb = _workbook(tmp_path / "form.xlsx", {"Form": [["Bank Name", "X"]]})
    doc = _pdf(tmp_path, "Bank confirmation letter\nIBAN DE89 3704 0044 0532 0130 00")
    res = run_check(doc, "bank", use_vision=False, engine="deterministic",
                    enforce_approvals=False, template_path=wb)
    assert any(f.rule_id == "TPL-014" for f in res.findings)
    assert res.verdict == "NEED_MANUAL_REVIEW"


def test_template_tin_is_secret(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    wb = _workbook(tmp_path / "form.xlsx", {"Form": [
        ["Bank Name", "Commerzbank AG"],
        ["IBAN Number", "DE89 3704 0044 0532 0130 00"],
        ["Tax Number 1", "000-11-2222"],
    ]})
    doc = _pdf(tmp_path, "Bank confirmation letter\n"
                         "IBAN: DE89 3704 0044 0532 0130 00")
    res = run_check(doc, "bank", use_vision=False, engine="deterministic",
                    enforce_approvals=False, template_path=wb)
    blob = (config.RUNS_DIR / res.run_id / "template_compare.json").read_text()
    assert "000-11-2222" not in blob
    md = (config.RUNS_DIR / res.run_id / "reasoning.md").read_text()
    assert "000-11-2222" not in md


REAL_TEMPLATE = "/Users/egor/Downloads/Americas MacroEnabled MDM_SETH_DEVRIES_HCP.xlsm"


@pytest.mark.skipif(not __import__("pathlib").Path(REAL_TEMPLATE).exists(),
                    reason="real Americas template not on this machine")
def test_real_americas_template_parses():
    from pathlib import Path
    fields, prov, hits = template_form.parse(Path(REAL_TEMPLATE))
    assert hits >= 6
    assert fields["bank_name"].upper().startswith("CHASE")
    assert fields["bank_key"].endswith("326")
    assert fields["bank_account"]
    assert template_form.looks_like_request_form(Path(REAL_TEMPLATE))
