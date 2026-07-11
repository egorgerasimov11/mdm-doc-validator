"""address.py — SOURCE_ADDRNUMBER="1", region via address-validator, and the
local-script international ADRC version."""
from __future__ import annotations

import openpyxl

from mdmdoc.consolidation import address


def test_assign_source_addrnumber():
    rows = {sh: [{"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "NEW_20260711_01"}]
            for sh in ("ADRC - Address", "ADR2 - Phone", "ADR6 - E-Mail")}
    address.assign_source_addrnumber(rows)
    for sh in rows:
        assert rows[sh][0]["SOURCE_ADDRNUMBER"] == "1"       # non-initial address #
        assert rows[sh][0]["SOURCE_ID"] == "N1"              # vendor link untouched


def test_fill_region_from_resolver(monkeypatch):
    monkeypatch.setattr(address, "_resolve_region", lambda c, city, postal, warnings_out=None: "010")
    rows = {
        "LFA1 - Supplier General": [{"LAND1": "CN", "ORT01": "BEIJING"}],
        "ADRC - Address": [{"COUNTRY": "CN", "CITY1": "BEIJING",
                            "POST_CODE1": "100025"}],
    }
    address.fill_region(rows)
    assert rows["LFA1 - Supplier General"][0]["REGIO"] == "010"
    assert rows["ADRC - Address"][0]["REGION"] == "010"


def test_fill_region_keeps_form_supplied(monkeypatch):
    monkeypatch.setattr(address, "_resolve_region",
                        lambda *a: (_ for _ in ()).throw(AssertionError("must not call")))
    rows = {"LFA1 - Supplier General": [{"LAND1": "US", "REGIO": "TX"}],
            "ADRC - Address": [{"COUNTRY": "US", "REGION": "TX"}]}
    address.fill_region(rows)                                # form value kept, resolver skipped
    assert rows["LFA1 - Supplier General"][0]["REGIO"] == "TX"


def test_fill_region_unresolved_warns(monkeypatch):
    monkeypatch.setattr(address, "_resolve_region", lambda c, city, postal, warnings_out=None: None)
    warns: list = []
    rows = {"LFA1 - Supplier General": [{"LAND1": "ZZ", "ORT01": "NOWHERE"}],
            "ADRC - Address": [{"COUNTRY": "ZZ", "CITY1": "NOWHERE"}]}
    address.fill_region(rows, warns)
    assert not rows["ADRC - Address"][0].get("REGION")
    assert warns and "REGION" in warns[0]


def _form_with_intl_block(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2. Vendor Details"
    ws["I13"] = "International Address"
    ws["I14"] = "Name 1"
    ws["J14"] = "上海市对外服务北京有限公司"
    ws["I20"] = "Building Number/Street"
    ws["J20"] = "北京市东城区青龙胡同甲1号"
    ws["I23"] = "City"
    ws["J23"] = "北京"
    wb.save(path)
    wb.close()
    return path


def test_add_international_version_cn(tmp_path):
    form = _form_with_intl_block(tmp_path / "f.xlsx")
    rows = {"ADRC - Address": [{"SOURCE_ID": "N1", "SOURCE_ADDRNUMBER": "1",
                               "COUNTRY": "CN", "NAME1": "SHANGHAI FOREIGN SERVICE",
                               "STREET": "ROOM 416", "CITY1": "BEIJING",
                               "REGION": "010", "DATE_FROM": "20260711",
                               "NAME2": "STALE"}]}
    address.add_international_version(rows, form)
    adrc = rows["ADRC - Address"]
    assert len(adrc) == 2
    assert not adrc[0].get("NATION")                         # romanized row (row0) untouched
    r1 = adrc[1]
    assert r1["NATION"] == "C"
    assert r1["NAME1"] == "上海市对外服务北京有限公司"
    assert r1["CITY1"] == "北京"
    assert "NAME2" not in r1                                 # stale English continuation cleared
    assert r1["SOURCE_ADDRNUMBER"] == "1" and r1["REGION"] == "010"  # inherited


def test_add_international_version_skips_non_cjk(tmp_path):
    form = _form_with_intl_block(tmp_path / "f.xlsx")
    rows = {"ADRC - Address": [{"COUNTRY": "US", "NAME1": "ACME"}]}
    address.add_international_version(rows, form)
    assert len(rows["ADRC - Address"]) == 1                  # US has no local-script version


def test_add_international_version_no_block(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "2. Vendor Details"
    p = tmp_path / "f.xlsx"
    wb.save(p)
    wb.close()
    rows = {"ADRC - Address": [{"COUNTRY": "CN", "NAME1": "X"}]}
    address.add_international_version(rows, p)
    assert len(rows["ADRC - Address"]) == 1                  # no block → no 2nd row


def test_fill_phone_country():
    rows = {"LFA1 - Supplier General": [{"LAND1": "CN"}],
            "ADR2 - Phone": [{"SOURCE_ID": "N1", "TEL_NUMBER": "010-1"}]}
    address.fill_phone_country(rows)                          # columns=None
    assert rows["ADR2 - Phone"][0]["COUNTRY"] == "CN"


def _form_with_region(path, region):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2. Vendor Details"
    ws["D24"] = "State/Province/Region"
    ws["E24"] = region
    wb.save(path)
    wb.close()
    return path


def test_fill_district_parses_from_region(tmp_path):
    form = _form_with_region(tmp_path / "f.xlsx", "CHAOYANG DISTRICT,BEIJING,CHINA")
    rows = {"ADRC - Address": [{"CITY1": "BEIJING", "COUNTRY": "CN", "REGION": "010"}],
            "LFA1 - Supplier General": [{"ORT01": "BEIJING"}]}
    address.fill_district(rows, form)                        # columns=None
    assert rows["ADRC - Address"][0]["CITY2"] == "CHAOYANG DISTRICT"
    assert rows["LFA1 - Supplier General"][0]["ORT02"] == "CHAOYANG DISTRICT"


def test_fill_district_single_region_skips(tmp_path):
    form = _form_with_region(tmp_path / "f.xlsx", "MI")       # a clean region code, not a district
    rows = {"ADRC - Address": [{"CITY1": "GRAND RAPIDS", "COUNTRY": "US", "REGION": "MI"}]}
    address.fill_district(rows, form)
    assert not rows["ADRC - Address"][0].get("CITY2")
