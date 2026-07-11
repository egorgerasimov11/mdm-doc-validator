"""template_io.BPTemplate — the append-only engine (no converter needed)."""
from __future__ import annotations

import openpyxl
import pytest

from consolidation_helpers import cell, make_template, prefill_vendor
from mdmdoc.consolidation.template_io import (
    BPTemplate,
    BPTemplateError,
    snapshot_workbook,
)

ROWS_ONE = {
    "LFA1 - Supplier General": [
        {"SOURCE_ID": "NEW_1", "NAME1": "ALPHA LLC", "LAND1": "US"}],
    "BUT000 - General": [{"SOURCE_ID": "NEW_1", "NAME_ORG1": "ALPHA LLC"}],
    "LFB1 - Company Code (Supplier)": [
        {"SOURCE_ID": "NEW_1", "BUKRS": "0601", "ZTERM": "Z000"},
        {"SOURCE_ID": "NEW_1", "BUKRS": "0432", "ZTERM": "Z000"}],
}


def _tpl(tmp_path, name="t.xlsx"):
    return BPTemplate(make_template(tmp_path / name))


class TestAppend:
    def test_empty_template_starts_row3(self, tmp_path):
        tpl = _tpl(tmp_path)
        plan = tpl.append_rows(ROWS_ONE)
        out = tpl.save_to(tmp_path / "out.xlsx")
        tpl.close()
        assert min(c["row"] for c in plan) == 3
        wb = openpyxl.load_workbook(out)
        assert cell(wb, "LFA1 - Supplier General", 3, "NAME1") == "ALPHA LLC"
        assert cell(wb, "LFB1 - Company Code (Supplier)", 3, "BUKRS") == "0601"
        assert cell(wb, "LFB1 - Company Code (Supplier)", 4, "BUKRS") == "0432"
        wb.close()

    def test_prefilled_template_appends_after_existing(self, tmp_path):
        path = prefill_vendor(make_template(tmp_path / "t.xlsx"))
        before = snapshot_workbook(openpyxl.load_workbook(path))
        tpl = BPTemplate(path)
        plan = tpl.append_rows(ROWS_ONE)
        out = tpl.save_to(tmp_path / "out.xlsx")
        tpl.close()
        lfa1_rows = [c["row"] for c in plan
                     if c["sheet"] == "LFA1 - Supplier General"]
        assert min(lfa1_rows) == 4  # row 3 is the pre-existing vendor
        wb = openpyxl.load_workbook(out)
        # pre-existing row untouched, byte for byte
        assert cell(wb, "LFA1 - Supplier General", 3, "NAME1") == "EXISTING VENDOR GMBH"
        after = snapshot_workbook(wb)
        wb.close()
        for sheet, pre in before.items():
            import mdmdoc.consolidation.template_io as tio
            wb2 = openpyxl.load_workbook(out)
            assert tio.sheet_value_hash(wb2[sheet], pre["rows"]) == pre["sha"], sheet
            wb2.close()

    def test_formatting_ghost_rows_ignored(self, tmp_path):
        path = make_template(tmp_path / "t.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["LFA1 - Supplier General"]
        for r in range(3, 40):  # style-only rows: no values
            ws.cell(row=r, column=1).fill = openpyxl.styles.PatternFill(
                "solid", fgColor="FFFF00")
        wb.save(path)
        wb.close()
        tpl = BPTemplate(path)
        assert tpl.last_used_row("LFA1 - Supplier General") == 2
        plan = tpl.append_rows(ROWS_ONE)
        tpl.close()
        assert min(c["row"] for c in plan
                   if c["sheet"] == "LFA1 - Supplier General") == 3

    def test_two_vendors_disjoint_ranges(self, tmp_path):
        tpl = _tpl(tmp_path)
        p1 = tpl.append_rows(ROWS_ONE)
        rows2 = {s: [dict(r, SOURCE_ID="NEW_2") for r in rows]
                 for s, rows in ROWS_ONE.items()}
        p2 = tpl.append_rows(rows2)
        tpl.close()
        for sheet in ROWS_ONE:
            r1 = {c["row"] for c in p1 if c["sheet"] == sheet}
            r2 = {c["row"] for c in p2 if c["sheet"] == sheet}
            assert r1 and r2 and not (r1 & r2)
            assert min(r2) == max(r1) + 1

    def test_gap_row_between_vendors_never_overwrites_below(self, tmp_path):
        path = prefill_vendor(make_template(tmp_path / "t.xlsx"))
        wb = openpyxl.load_workbook(path)
        ws = wb["LFA1 - Supplier General"]
        hdr = {str(c.value).strip(): c.column for c in ws[2] if c.value}
        ws.cell(row=5, column=hdr["NAME1"], value="BELOW THE GAP AG")  # row 4 blank
        wb.save(path)
        wb.close()
        tpl = BPTemplate(path)
        plan = tpl.append_rows(ROWS_ONE)
        tpl.close()
        lfa1 = [c["row"] for c in plan if c["sheet"] == "LFA1 - Supplier General"]
        assert min(lfa1) == 6  # after the LAST used row, not the first blank

    def test_numeric_values_stringified(self, tmp_path):
        tpl = _tpl(tmp_path)
        plan = tpl.append_rows({"BUT0BK - Bank Account": [
            {"SOURCE_ID": "NEW_1", "BANKL": 72000326.0, "BANKN": 683661570}]})
        out = tpl.save_to(tmp_path / "out.xlsx")
        tpl.close()
        by_tech = {c["tech"]: c["value"] for c in plan}
        assert by_tech["BANKL"] == "72000326"      # no trailing .0
        assert by_tech["BANKN"] == "683661570"
        wb = openpyxl.load_workbook(out)
        assert cell(wb, "BUT0BK - Bank Account", 3, "BANKL") == "72000326"
        wb.close()


class TestGuards:
    def test_foreign_xlsx_fails_loud(self, tmp_path):
        p = tmp_path / "foreign.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "just a spreadsheet"
        wb.save(p)
        wb.close()
        tpl = BPTemplate(p)
        with pytest.raises(BPTemplateError):
            tpl.validate()
        tpl.close()

    def test_missing_required_sheet_fails_loud(self, tmp_path):
        from consolidation_helpers import TEMPLATE_SHEETS
        sheets = {k: v for k, v in TEMPLATE_SHEETS.items()
                  if k != "BUT0BK - Bank Account"}
        tpl = BPTemplate(make_template(tmp_path / "t.xlsx", sheets))
        with pytest.raises(BPTemplateError, match="BUT0BK"):
            tpl.validate()
        tpl.close()

    def test_comment_column_never_written(self, tmp_path):
        tpl = _tpl(tmp_path)
        with pytest.raises(BPTemplateError, match="_COMMENT"):
            tpl.append_rows({"LFA1 - Supplier General": [
                {"_COMMENT": "boom", "NAME1": "X"}]})
        tpl.close()

    def test_unknown_column_fails_loud_before_any_write(self, tmp_path):
        tpl = _tpl(tmp_path)
        pre = tpl.snapshot()
        with pytest.raises(BPTemplateError, match="NO_SUCH_COL"):
            tpl.append_rows({"LFA1 - Supplier General": [
                {"NAME1": "X", "NO_SUCH_COL": "Y"}]})
        assert tpl.snapshot() == pre  # nothing was written
        tpl.close()

    def test_action_code_only_where_column_exists(self, tmp_path):
        # BUT000 has no _ACTION_CODE column in the real template — a row
        # carrying it must fail loud there but pass on LFB1.
        tpl = _tpl(tmp_path)
        with pytest.raises(BPTemplateError, match="_ACTION_CODE"):
            tpl.append_rows({"BUT000 - General": [
                {"NAME_ORG1": "X", "_ACTION_CODE": "I"}]})
        plan = tpl.append_rows({"LFB1 - Company Code (Supplier)": [
            {"BUKRS": "0601", "_ACTION_CODE": "I"}]})
        tpl.close()
        assert {"_ACTION_CODE"} <= {c["tech"] for c in plan}

    def test_formula_values_rejected(self, tmp_path):
        # '='-prefixed strings become Excel FORMULAS in openpyxl: the verify
        # passes would compare the formula text and pass while Excel/SAP
        # evaluates something else (formula injection from form values)
        tpl = _tpl(tmp_path)
        with pytest.raises(BPTemplateError, match="formula"):
            tpl.append_rows({"LFA1 - Supplier General": [
                {"NAME1": "=HYPERLINK(\"http://evil\")", "SOURCE_ID": "N1"}]})
        tpl.close()

    def test_refuses_overwriting_the_upload(self, tmp_path):
        path = make_template(tmp_path / "t.xlsx")
        tpl = BPTemplate(path)
        tpl.append_rows(ROWS_ONE)
        with pytest.raises(BPTemplateError, match="refusing"):
            tpl.save_to(path)
        tpl.close()


class TestHeaderDetection:
    def test_sap_example_layout_row6(self, tmp_path):
        """SUPPLIER.xlsx-style layout: legend rows 1-5, SOURCE_ID header row 6."""
        p = tmp_path / "example.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("LFA1 - Supplier General")
        ws["A1"] = "Delete or hide this area when using this file for importing data."
        ws["A3"], ws["B3"] = "Table", "LFA1"
        ws["A4"], ws["B4"] = "Description", "Supplier General"
        for i, t in enumerate(["SOURCE_ID", "LIFNR", "NAME1", "LAND1"], start=1):
            ws.cell(row=5, column=i, value=f"desc {t}")
            ws.cell(row=6, column=i, value=t)
        wb.save(p)
        wb.close()
        tpl = BPTemplate(p)
        assert tpl.header_row("LFA1 - Supplier General") == 6
        plan = tpl.append_rows({"LFA1 - Supplier General": [
            {"SOURCE_ID": "NEW_1", "NAME1": "ALPHA"}]})
        tpl.close()
        assert min(c["row"] for c in plan) == 7

    def test_existing_ids_and_names(self, tmp_path):
        path = prefill_vendor(make_template(tmp_path / "t.xlsx"),
                              source_id="OLD_9", name="Existing  Vendor GmbH")
        tpl = BPTemplate(path)
        assert "OLD_9" in tpl.existing_source_ids()
        assert "EXISTING VENDOR GMBH" in tpl.existing_vendor_names()
        tpl.close()
