"""Attach a document to a form vendor → merge, quiz, consolidate (API)."""
from __future__ import annotations

import io

from consolidation_helpers import (
    bank_run_fields,
    make_americas_form,
    make_template,
    needs_converter,
    prefill_vendor,
    write_run,
)
from mdmdoc import config

pytestmark = needs_converter


def _client(monkeypatch, tmp_path):
    from fastapi.testclient import TestClient

    from mdmdoc import runstore
    from mdmdoc.server.app import create_app
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(config, "INBOX_DIR", tmp_path / "inbox")
    monkeypatch.setattr(runstore, "RUNS_DIR", tmp_path / "runs", raising=False)
    monkeypatch.setenv("MDMDOC_BANK_VALUES", "full")
    monkeypatch.setenv("MDMDOC_TIN_VALUES", "full")
    return TestClient(create_app("full"))


def _new_case(c):
    r = c.post("/ui/consolidation/new", follow_redirects=False)
    return r.headers["location"].rstrip("/").rsplit("/", 1)[-1]


def _up(c, case, route, path, name=None, data=None):
    return c.post(f"/ui/consolidation/{case}/{route}",
                  files={"file": (name or path.name, io.BytesIO(path.read_bytes()),
                                  "application/octet-stream")},
                  data=data or {}, follow_redirects=False)


def _setup(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    write_run(config.RUNS_DIR, "bankdoc", "bank", "ACCEPT", bank_run_fields())
    form = make_americas_form(tmp_path / "f.xlsm")
    tpl = make_template(tmp_path / "tpl.xlsx")
    case = _new_case(client)
    _up(client, case, "template", tpl)
    _up(client, case, "extract", form)
    return client, case


def test_attach_shows_quiz_and_crosscheck(monkeypatch, tmp_path):
    client, case = _setup(monkeypatch, tmp_path)
    r = client.post(f"/ui/consolidation/{case}/vendor/v01/attach-run",
                    data={"run_id": "bankdoc"}, follow_redirects=False)
    assert r.status_code == 303
    page = client.get(f"/ui/consolidation/{case}").text
    assert "choice_BANKL" in page and "072000326" in page and "021000021" in page
    assert "683661570" not in page  # account masked everywhere in the page
    # candidates route reachable
    cands = client.get("/ui/consolidation/attach/candidates").json()
    assert any(c["run_id"] == "bankdoc" for c in cands["candidates"])


def test_consolidate_blocked_until_quiz_resolved(monkeypatch, tmp_path):
    client, case = _setup(monkeypatch, tmp_path)
    client.post(f"/ui/consolidation/{case}/vendor/v01/attach-run",
                data={"run_id": "bankdoc"})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"})
    assert "resolve the field choice" in r.text
    assert client.get(f"/ui/consolidation/{case}/download").status_code == 409


def test_resolve_then_consolidate_writes_chosen_routing(monkeypatch, tmp_path):
    client, case = _setup(monkeypatch, tmp_path)
    client.post(f"/ui/consolidation/{case}/vendor/v01/attach-run",
                data={"run_id": "bankdoc"})
    client.post(f"/ui/consolidation/{case}/vendor/v01/resolve",
                data={"choice_BANKL": "doc_wires"})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    dl = client.get(f"/ui/consolidation/{case}/download")
    assert dl.status_code == 200
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["BUT0BK - Bank Account"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    assert ws.cell(row=3, column=h["BANKL"]).value == "021000021"   # the wires pick
    assert ws.cell(row=3, column=h["BANKN"]).value == "000661570"   # account
    wb.close()
    # leak gate: full account only in the workbook, not in case JSON
    case_dir = config.RUNS_DIR / "consolidation" / case
    for p in case_dir.rglob("*.json"):
        assert "000661570" not in p.read_text(encoding="utf-8"), p.name


def test_attach_default_doc_wins_and_fixes_aba(monkeypatch, tmp_path):
    # a single document routing that differs from the form -> ambiguous, but
    # picking the document (default) writes the corrected 9-digit ABA
    client = _client(monkeypatch, tmp_path)
    write_run(config.RUNS_DIR, "d", "bank", "ACCEPT",
              bank_run_fields(routing_ach="072000326", routing_wires=None))
    form = make_americas_form(tmp_path / "f.xlsm")   # form ABA 71000013 (8-digit)
    tpl = make_template(tmp_path / "tpl.xlsx")
    case = _new_case(client)
    _up(client, case, "template", tpl)
    _up(client, case, "extract", form)
    client.post(f"/ui/consolidation/{case}/vendor/v01/attach-run", data={"run_id": "d"})
    client.post(f"/ui/consolidation/{case}/vendor/v01/resolve",
                data={"choice_BANKL": "doc_ach"})
    r = client.post(f"/ui/consolidation/{case}/consolidate",
                    data={"confirm_warnings": "on"}, follow_redirects=False)
    assert r.status_code == 303
    dl = client.get(f"/ui/consolidation/{case}/download")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["BUT0BK - Bank Account"]
    h = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    assert ws.cell(row=3, column=h["BANKL"]).value == "072000326"  # corrected
    wb.close()


def test_detach_reverts_to_form_only(monkeypatch, tmp_path):
    client, case = _setup(monkeypatch, tmp_path)
    client.post(f"/ui/consolidation/{case}/vendor/v01/attach-run", data={"run_id": "bankdoc"})
    r = client.post(f"/ui/consolidation/{case}/vendor/v01/detach", follow_redirects=False)
    assert r.status_code == 303
    page = client.get(f"/ui/consolidation/{case}").text
    assert "choice_BANKL" not in page  # quiz gone


def test_attach_refused_on_reject_run(monkeypatch, tmp_path):
    client, case = _setup(monkeypatch, tmp_path)
    write_run(config.RUNS_DIR, "bad", "bank", "REJECT", bank_run_fields())
    r = client.post(f"/ui/consolidation/{case}/vendor/v01/attach-run",
                    data={"run_id": "bad"})
    assert "REJECT" in r.text


def test_clear_checkbox_via_api(monkeypatch, tmp_path):
    client = _client(monkeypatch, tmp_path)
    tpl = prefill_vendor(make_template(tmp_path / "tpl.xlsx"))
    case = _new_case(client)
    r = _up(client, case, "template", tpl, data={"clear_existing": "on"})
    assert r.status_code == 303
    page = client.get(f"/ui/consolidation/{case}").text
    assert "cleared" in page and "existing vendors: 0" in page
