"""Shared builders for the consolidation test suite (not collected by pytest).

Synthesizes (a) a mini BusinessPartnerTemplate mirroring the REAL layout —
descriptions in row 1, technical names in row 2, data from row 3, and BUT000
deliberately WITHOUT an _ACTION_CODE column (the real SAP-delivered template
lacks it there) — and (b) an Americas vendor request form shaped like the
SETH_DEVRIES packet but with fake PII only.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from mdmdoc.consolidation import available

needs_converter = pytest.mark.skipif(
    not available(),
    reason="sap_vendor_autoload not installed — run: "
           "uv run --with-editable <sap-vendor-autoload> pytest")

TEMPLATE_SHEETS: dict[str, list[str]] = {
    "! Read Me !": ["_COMMENT"],
    "BUT000 - General": [  # NB: no _ACTION_CODE, like the real template
        "_COMMENT", "SOURCE_ID", "PARTNER", "TYPE", "BU_GROUP",
        "NAME_ORG1", "NAME_ORG2", "BU_SORT1", "BU_SORT2", "BU_LANGU"],
    "BUT100 - Role": ["_COMMENT", "_ACTION_CODE", "SOURCE_ID", "PARTNER", "RLTYP"],
    "LFA1 - Supplier General": [
        "_COMMENT", "_ACTION_CODE", "SOURCE_ID", "LIFNR", "KTOKK", "SORTL",
        "NAME1", "NAME2", "STRAS", "ORT02", "PSTLZ", "ORT01", "LAND1",
        "REGIO", "SPRAS", "TELF1", "TELFX", "STCD1", "STCD2", "STCD3",
        "STCD4", "STCEG"],
    "LFB1 - Company Code (Supplier)": [
        "_COMMENT", "_ACTION_CODE", "SOURCE_ID", "BUKRS", "ASSIGNMENT_ID",
        "ZTERM", "ZWELS", "QSREC", "AKONT", "ZUAWA", "FDGRV", "EIKTO"],
    "LFM1 - Purchasing Org Data": [
        "_COMMENT", "_ACTION_CODE", "SOURCE_ID", "EKORG", "ASSIGNMENT_ID",
        "WAERS", "ZTERM"],
    "ADRC - Address": [
        "_COMMENT", "_ACTION_CODE", "SOURCE_ID", "SOURCE_ADDRNUMBER",
        "NAME1", "NAME2", "SORT2", "STREET", "STR_SUPPL2", "CITY2",
        "POST_CODE1", "CITY1", "COUNTRY", "REGION", "LANGU",
        "TEL_NUMBER", "FAX_NUMBER"],
    "ADR2 - Phone": ["_COMMENT", "_ACTION_CODE", "SOURCE_ID",
                     "SOURCE_ADDRNUMBER", "TEL_NUMBER", "FAX_NUMBER"],
    "ADR6 - E-Mail": ["_COMMENT", "_ACTION_CODE", "SOURCE_ID",
                      "SOURCE_ADDRNUMBER", "SMTP_ADDR"],
    "BUT0BK - Bank Account": [
        "_COMMENT", "_ACTION_CODE", "SOURCE_ID", "PARTNER", "BKVID",
        "BANKS", "BANKL", "BANKN", "BKONT", "BKREF", "KOINH", "IBAN"],
    "BUT0ID - Identifier": ["_COMMENT", "_ACTION_CODE", "SOURCE_ID",
                            "PARTNER", "TYPE", "IDNUMBER"],
    "DFKKBPTAXNUM - Tax Number": ["_COMMENT", "_ACTION_CODE", "SOURCE_ID",
                                  "TAXTYPE", "PARTNER", "TAXNUM", "TAXNUMXL"],
}

# fake vendor: no real PII. Routing 71000013 zero-pads to the checksum-valid
# 071000013 (ABA 3-7-1), reproducing the real form's lost-leading-zero case.
FAKE = {
    "vendor_name": "SETH FAKESON",
    "street": "1602 FAKE RD SE",
    "postal": "49506",
    "city": "GRAND RAPIDS",
    "region": "MI",
    "country": "USA",
    "language": "ENG",
    "email": "seth.fakeson@example.org",
    "bank_name": "FAKE BANK",
    "bank_country": "USA",
    "bank_key8": "71000013",
    "bank_account": "000661570",
    "ssn": "000-04-2016",
}


def make_template(path, sheets: dict | None = None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, headers in (sheets or TEMPLATE_SHEETS).items():
        ws = wb.create_sheet(name)
        for idx, tech in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=f"desc {tech}")
            ws.cell(row=2, column=idx, value=tech)
    wb.save(path)
    wb.close()
    return path


def prefill_vendor(path, source_id="OLD_1", name="EXISTING VENDOR GMBH"):
    """Append a minimal pre-existing vendor so appends must land BELOW it."""
    wb = openpyxl.load_workbook(path)
    for sheet, values in {
        "LFA1 - Supplier General": {"SOURCE_ID": source_id, "NAME1": name,
                                    "LAND1": "DE"},
        "BUT000 - General": {"SOURCE_ID": source_id, "NAME_ORG1": name},
        "LFB1 - Company Code (Supplier)": {"SOURCE_ID": source_id,
                                           "BUKRS": "0100"},
    }.items():
        ws = wb[sheet]
        cols = {str(c.value).strip(): c.column for c in ws[2] if c.value}
        row = 3
        while any(ws.cell(row=row, column=c).value for c in cols.values()):
            row += 1
        for tech, v in values.items():
            ws.cell(row=row, column=cols[tech], value=v)
    wb.save(path)
    wb.close()
    return path


def _labeled(ws, label_cell, value_cell, label, value):
    ws[label_cell] = label
    ws[value_cell] = value


def make_americas_form(path, **overrides):
    vals = {**FAKE, **overrides}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1. General Info"
    details = wb.create_sheet("2. Vendor Details")
    wb.create_sheet("Bank Info Validation")
    wb.create_sheet("DROPDOWN VALUES NA")

    general = wb["1. General Info"]
    for row, label, value in [
        (7, "Requestor Name", "FAKE REQUESTOR"),
        (9, "Request Type", "New"),
        (10, "Vendor Number", vals.get("vendor_number", "(Required if existing)")),
        (11, "Vendor Name", vals["vendor_name"]),
    ]:
        _labeled(general, f"B{row}", f"C{row}", label, value)

    for row, label, value in [
        (8, "Company Code", "FakeCo USA, Inc. 0601, FakeCo, Inc. 0432"),
        (9, "Purchasing Organization", "Houston U001, Arvada U002"),
        (10, "Account Group", "ZKTV - Third Vendor"),
        (15, "Name 1", vals["vendor_name"]),
        (17, "Search Term 1", vals["vendor_name"]),
        (18, "Search Term 2", "HCE"),
        (19, "Building Number/Street", vals["street"]),
        (22, "Postal Code", vals["postal"]),
        (23, "City", vals["city"]),
        (24, "State/Province/Region", vals["region"]),
        (25, "Country", vals["country"]),
        (26, "Language", vals["language"]),
        (28, "E-mail", vals["email"]),
        (41, "Bank Name", vals["bank_name"]),
        (42, "Bank Country", vals["bank_country"]),
        (43, "Bank Key (ABA Routing Code)", vals["bank_key8"]),
        (44, "Bank Account Number", vals["bank_account"]),
        (47, "Control Key", "(Required for Colombia)"),
        (48, "Reference Details", "(Required for Colombia)"),
        (51, "Tax Number 1", vals["ssn"]),
        (52, "Tax Number 2", "(Required for US companies)"),
        (57, "Order Currency", "USD"),
        (58, "Payment Terms", "Z000 - Payable Immediately"),
        (64, "Recon Account", "B344001000"),
        (65, "Sort Key", "Z01"),
        (66, "Cash Management Group", "VTP_DOM_BI"),
        (67, "Payment Method", "A - ACH"),
        (68, "Recipient Type", "1 - Individual"),
    ]:
        _labeled(details, f"D{row}", f"E{row}", label, value)

    wb.save(path)
    wb.close()
    return path


import json as _json


def write_run(runs_dir, run_id, doc_class, verdict, fields, *, file_name=None,
              template_path=None):
    """Synthesize a run dir (meta/report/extraction) for merge/attach tests.
    `fields` sensitive entries are {"value", "present", "masked"} dicts; plain
    strings for non-sensitive fields."""
    d = Path(runs_dir) / run_id
    d.mkdir(parents=True, exist_ok=True)
    meta = {"run_id": run_id, "file_name": file_name or f"{run_id}.pdf",
            "doc_class": doc_class, "test": False}   # an operator document
    if template_path:
        meta["template_path"] = str(template_path)
    (d / "meta.json").write_text(_json.dumps(meta))
    (d / "report.json").write_text(_json.dumps(
        {"schema": "mdmdoc.v1", "doc_class": doc_class, "verdict": verdict}))
    (d / "extraction.json").write_text(_json.dumps(
        {"doc_class": doc_class, "fields": fields, "provenance": {}}))
    return run_id


def bank_run_fields(routing_ach="072000326", routing_wires="021000021",
                    account="000661570", holder="SETH FAKESON", country="US"):
    # account defaults to the form's FAKE bank_account so only the ROUTING is
    # ambiguous (the real SETH scenario: account agrees, only the ABA differs)
    f = {"account_holder": holder, "bank_country": country}
    if routing_ach:
        f["routing_aba"] = {"value": routing_ach, "present": True,
                            "masked": routing_ach[:4] + "*"}
    if routing_wires:
        f["routing_aba_wires"] = {"value": routing_wires, "present": True,
                                  "masked": routing_wires[:4] + "*"}
    if account:
        f["account_number"] = {"value": account, "present": True,
                               "masked": "***" + account[-3:]}
    return f


def headers_of(ws, header_row=2):
    return {str(c.value).strip(): c.column for c in ws[header_row] if c.value}


def cell(wb, sheet, row, tech, header_row=2):
    ws = wb[sheet]
    return ws.cell(row=row, column=headers_of(ws, header_row)[tech]).value
