#!/usr/bin/env python3
"""Generate the canonical bulk templates (templates/bulk/*.xlsx) — V5.

Deterministic: same inputs -> the same three workbooks. Each template carries
  Data     — the STRICT canonical headers bulk/reader.py parses (row 1);
  Examples — 2-3 fully INVENTED rows showing how to fill it;
  README   — the per-case guideline (columns, where the data lives in SAP,
             what every bucket and BULK-### rule means).
Run: uv run python tools/gen_bulk_templates.py
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

OUT = ROOT / "templates" / "bulk"

BUCKET_LEGEND = [
    ("VALID", "every applicable deterministic check passed"),
    ("SUSPICIOUS", "formally passing but needs a human look (duplicates, "
                   "registry miss, empty-but-expected)"),
    ("INVALID", "a deterministic check FAILED — checksum/format/wrong-country; "
                "mathematically or structurally wrong"),
    ("SKIPPED", "cannot be judged (masked in the export, empty row)"),
]

CASES = {
    "banking_template.xlsx": {
        "title": "Bulk banking validation — canonical template",
        "headers": ["Business Partner", "Bank Country", "Bank Key",
                    "Bank Account", "IBAN", "SWIFT", "Bank Control Key",
                    "Account Holder", "Account Name"],
        "examples": [
            ["1000001", "US", "021000021", "12345678", "", "", "01",
             "NORTHWIND TRADING LLC", ""],
            ["1000002", "DE", "10070000", "", "DE89370400440532013000",
             "DEUTDEBBXXX", "", "Altair GmbH", ""],
            ["1000003", "GB", "200000", "55779911", "GB29NWBK60161331926819",
             "NWBKGB2LXXX", "", "Kestrel Ltd", ""],
        ],
        "readme": [
            "WHERE THE DATA LIVES IN SAP: table BUT0BK (BP bank details) —",
            "SE16N export works directly too (its headers are recognized).",
            "Business Partner = BUT0BK-PARTNER, Bank Country = BANKS,",
            "Bank Key = BANKL (US: the 9-digit ABA routing number),",
            "Bank Account = BANKN, Control Key = BKONT, Holder = KOINH.",
            "",
            "CHECKS (rule ids cited in every verdict):",
            "BULK-B01 US bank key must be exactly 9 digits (SWIFT in the key",
            "         field is a classic mis-entry and is named as such).",
            "BULK-B02 ABA 3-7-1 mod-10 checksum — a failing routing number",
            "         mathematically cannot exist (ABA standard).",
            "BULK-B03 Fed routing prefix ranges: 00, 01-12, 21-32, 61-72, 80.",
            "BULK-B04 account all zeros = INVALID; under 4 significant digits",
            "         after zero-padding = SUSPICIOUS (leading zeros are SAP",
            "         padding, never a defect by themselves).",
            "BULK-B05 IBAN: ISO 13616 mod-97 checksum + national length.",
            "BULK-B06 IBAN country prefix vs Bank Country.",
            "BULK-B07 SWIFT/BIC shape (ISO 9362) + its country vs Bank Country.",
            "BULK-B08 control key convention (US: 01 checking / 02 savings).",
            "BULK-B09 the same account under several partners -> duplicate.",
            "BULK-B10 national key shapes: DE BLZ 8, GB sort 6, AU BSB 6,",
            "         IN IFSC AAAA0XXXXXX, CN CNAPS 12, JP 4+3, CA 9…",
            "BULK-B11 empty/masked account -> SKIPPED (not judged).",
            "BULK-B12 (web option ON) routing not listed in any of 3 live",
            "         public directories (usbanklocations/paymentlabs/wise)",
            "         -> SUSPICIOUS; found -> the bank name is cited.",
        ],
    },
    "tax_template.xlsx": {
        "title": "Bulk tax-number validation — canonical template",
        "headers": ["Business Partner", "Tax Number Category", "Tax Number",
                    "Tax Number Long", "Country"],
        "examples": [
            ["1000001", "US2", "12-3456789", "", "US"],
            ["1000002", "DE0", "DE811907980", "", "DE"],
            ["1000003", "IT0", "00743110157", "", "IT"],
        ],
        "readme": [
            "WHERE THE DATA LIVES IN SAP: BP tax numbers (DFKKBPTAXNUM /",
            "'Tax Numbers' facet) — an SE16N export works directly.",
            "Tax Number Category = SAP category (US1, DE0, IT0, …);",
            "Tax Number Long holds values over 20 chars when SAP split them.",
            "",
            "US DOCTRINE: the category digit (US0..US4) carries NO authoritative",
            "SSN/EIN mapping — every US value is judged by the NUMBER'S OWN",
            "STRUCTURE (valid if it passes ANY of SSN / ITIN / EIN rules).",
            "",
            "CHECKS:",
            "BULK-T01 category not in the catalog (rules/bulk_tax.yaml) ->",
            "         SUSPICIOUS, judged by shape only. Add local categories",
            "         to the catalog file.",
            "BULK-T02 value does not match the category's format/structure.",
            "BULK-T03 value is ANOTHER country's number: national prefix +",
            "         shape (+checksum where defined) proves it, e.g. a German",
            "         VAT DE1371… stored under a US category -> INVALID.",
            "BULK-T04 country checksum failed (DE ISO-7064, IT, FR key, ES",
            "         NIF/CIF letters, PL, BE, AT, GB 97/9755, CH, ABN, CNPJ,",
            "         CPF, CN USCC, IN GSTIN…).",
            "BULK-T05 masked value (XXXXXXX, ****1234) -> SKIPPED.",
            "BULK-T06 the same number under several partners -> duplicate.",
            "BULK-T07 empty value -> SKIPPED.",
            "BULK-T08 the Country column disagrees with the category country.",
        ],
    },
    "postal_region_template.xlsx": {
        "title": "Bulk postal/region validation — canonical template",
        "headers": ["Business Partner", "Country", "Region", "Postal Code",
                    "City"],
        "examples": [
            ["1000001", "US", "TX", "75201", "Dallas"],
            ["1000002", "DE", "BY", "80331", "München"],
            ["1000003", "CN", "010", "100020", "北京"],
        ],
        "readme": [
            "WHERE THE DATA LIVES IN SAP: BP addresses (ADRC / BUT020 facet).",
            "Country = ADRC-COUNTRY, Region = ADRC-REGION, Postal = PSTLZ.",
            "",
            "ATTACH THE REFERENCES to the run (Bulk page '+ reference'):",
            "  * T005S export (Country/Region Key | Region | Description) —",
            "    enables region MEMBERSHIP checks;",
            "  * T005U export (adds language texts) — optional, enriches",
            "    reasons with the region's description.",
            "Without T005S only postal formats are checked (noted in result).",
            "",
            "CHECKS:",
            "BULK-R01 country not ISO-mappable -> SUSPICIOUS.",
            "BULK-R02 region code absent from T005S for that country ->",
            "         INVALID (valid samples are cited).",
            "BULK-R03 region EMPTY though the country has regions in T005S:",
            "         region-required countries (US CA AU BR MX IN CN JP IT",
            "         ES) -> INVALID, others -> SUSPICIOUS.",
            "BULK-R04 postal code does not match the country's format",
            "         (rules/bulk_postal.yaml, ~85 countries).",
            "BULK-R05 a postal code in a country that has none (AE/HK/QA…).",
            "BULK-R06 placeholders ('Foreign', '99', 'XX', '000…').",
            "BULK-R07 empty row -> SKIPPED.",
        ],
    },
}


def build() -> list[Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    OUT.mkdir(parents=True, exist_ok=True)
    written = []
    for fname, spec in CASES.items():
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(spec["headers"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for i, h in enumerate(spec["headers"], start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)
        ws.freeze_panes = "A2"

        ex = wb.create_sheet("Examples")
        ex.append(spec["headers"])
        for c in ex[1]:
            c.font = Font(bold=True)
        for row in spec["examples"]:
            ex.append(row)
        ex.append([])
        ex.append(["^ INVENTED examples — delete this sheet before filling; "
                   "only the Data sheet is read."])

        rd = wb.create_sheet("README")
        rd.append([spec["title"]])
        rd["A1"].font = Font(bold=True, size=13)
        rd.append([])
        rd.append(["HOW TO USE: fill the Data sheet (or let the "
                   "mdmdoc-bulk-feed skill fill it from raw SAP exports), "
                   "then drop the file on the console's Bulk page or run "
                   "`mdmdoc bulk <file>`. Raw SE16N exports with descriptive "
                   "headers are recognized directly as well."])
        rd.append([])
        rd.append(["VERDICT BUCKETS:"])
        for b, meaning in BUCKET_LEGEND:
            rd.append([f"  {b}", meaning])
        rd.append([])
        for line in spec["readme"]:
            rd.append([line])
        rd.column_dimensions["A"].width = 100

        out = OUT / fname
        wb.save(out)
        written.append(out)
        print(f"wrote {out}")
    return written


if __name__ == "__main__":
    build()
