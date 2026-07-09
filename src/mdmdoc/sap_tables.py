#!/usr/bin/env python3
"""
sap_tables.py — read an SAP table EXPORT (SE16/SE16N .xlsx) and compare a
document against the stored Business Partner rows.

Two table kinds, detected by a header fingerprint (descriptive SE16N headers OR
raw technical field names both work):
  * BUT0BK — bank details (one row per bank account of a partner): maps to the
    same 14 SAP_KEYS the screenshot path produces, so sap_compare.compare() is
    reused UNCHANGED (bank comparison).
  * BUT000 — BP general data (name / category / blocks): a small dedicated
    comparison for W-9 / tax documents (name + individual-vs-entity + blocks).

Privacy: this is a LOCAL, deterministic reader (friendlier to the sealed BTP
image than the vision screenshot path). Only the SELECTED partner row's account
identifiers are registered as run secrets — never all 500 rows. Reverse-lookup
findings name PARTNER NUMBERS only, never account digits.
"""
from __future__ import annotations

import re

from .fields import Extraction, _norm_id, _norm_name, names_materially_equal
from .privacy import FIELD_KIND, display_value
from .rules.engine import Finding

# --- header aliasing: descriptive SE16N header OR technical name -> canon key ---
# canon keys are the technical BUT0BK/BUT000 field names.
_ALIASES: dict[str, str] = {}


def _alias(canon: str, *names: str) -> None:
    _ALIASES[canon.lower()] = canon
    for n in names:
        _ALIASES[_hkey(n)] = canon


def _hkey(s: str) -> str:
    """Normalize a header cell for matching: lowercase, collapse non-alphanumerics."""
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


# BUT0BK (bank details)
_alias("PARTNER", "Business Partner", "Partner")
_alias("BKVID", "Bank Details ID")
_alias("BANKS", "Bank Country Region", "Bank Country/Region", "Bank Country")
_alias("BANKL", "Bank Key")
_alias("BANKN", "Bank acct", "Bank Account Number", "Bank account")
_alias("BKONT", "Bank Control Key")
_alias("BKREF", "Reference Details")
_alias("KOINH", "Account Holder Name", "Account Holder")
_alias("ACCNAME", "Account Name")
_alias("IBAN", "IBAN")
_alias("BANKING_TYPE", "Bank Account Type")
# BUT000 (general data)
_alias("TYPE", "Partner Cat", "Partner Category")
_alias("BU_GROUP", "Grouping")
_alias("NAME_ORG1", "Name 1")
_alias("NAME_ORG2", "Name 2")
_alias("NAME_ORG3", "Name 3")
_alias("NAME_ORG4", "Name 4")
_alias("NAME_LAST", "Last Name")
_alias("NAME_FIRST", "First Name")
_alias("BU_FULLNAME", "Full Name")
_alias("BU_SORT1", "Search Term 1")
_alias("XBLCK", "Central Block")
_alias("XDELE", "Archiving Flag")
_alias("NATPERS", "Natural Person")
_alias("LEGAL_ENTITY", "Legal entity")

# BUT0BK is a per-account table → it has a Bank Details ID / Bank account column
# that BUT000 never carries. BUT000 (BP general data) carries name+category and,
# in a wide SE16N dump, ALSO a trailing Bank Country/Bank Key pair (BANKS/BANKL)
# — so BANKL alone must NOT trigger BUT0BK. Detect on the columns that are unique
# to each table.
_BUT0BK_MARKERS = {"BANKN", "BKVID"}
_BUT000_MARKERS = {"NAME_ORG1", "BU_FULLNAME"}


class SapTableError(ValueError):
    pass


def _cell_str(v) -> str:
    """A cell value as a clean string. openpyxl returns numeric cells as int/float
    — render without a trailing '.0' (SE16 exports usually keep IDs as text, so
    leading zeros survive; a numeric cell that already lost its leading zero at
    export time cannot be recovered here — the compare's zero-pad tolerance
    covers accounts)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "X" if v else ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load(path) -> tuple[str, list[dict]]:
    """(kind, rows) — kind in {'BUT0BK','BUT000'}; rows keyed by technical names.
    Reads the 'Data' sheet (or the first sheet); the first non-empty row that
    resolves >=3 known columns is the header row (SE16 exports carry a title/
    version preamble on some downloads)."""
    try:
        import openpyxl
    except ImportError as e:   # pragma: no cover
        raise SapTableError("openpyxl is required to read SAP table exports") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.worksheets[0]
    header_idx, col_map = None, {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        m = {}
        for c, cell in enumerate(row):
            canon = _ALIASES.get(_hkey(cell)) or _ALIASES.get(str(cell or "").lower())
            if canon and canon not in m:
                m[canon] = c
        if len(m) >= 3:
            header_idx, col_map = i, m
            break
    if header_idx is None:
        wb.close()
        raise SapTableError("not a recognized BUT0BK/BUT000 export (no known headers)")
    cols = set(col_map)
    has_bk = bool(_BUT0BK_MARKERS & cols)          # bank-details-specific columns
    has_bp = bool(_BUT000_MARKERS & cols) and "TYPE" in cols   # general-data-specific
    if has_bk and not has_bp:
        kind = "BUT0BK"
    elif has_bp and not has_bk:
        kind = "BUT000"
    elif has_bk:                                    # both present → bank-details wins
        kind = "BUT0BK"
    elif has_bp:
        kind = "BUT000"
    else:
        wb.close()
        raise SapTableError("known headers found but neither BUT0BK nor BUT000 markers")
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        d = {canon: _cell_str(row[c]) for canon, c in col_map.items() if c < len(row)}
        if any(d.values()):
            rows.append(d)
    wb.close()
    return kind, rows


# --- BUT0BK: row selection + reverse lookup + SAP_KEYS adapter ------------------
def to_sap_fields(row: dict) -> dict:
    """One BUT0BK row -> the 14-key SAP_KEYS dict compare() consumes."""
    return {
        "bank_details_id": row.get("BKVID", ""),
        "bank_account": row.get("BANKN", ""),
        "account_holder": row.get("KOINH", ""),
        "account_name": row.get("ACCNAME", ""),
        "iban": row.get("IBAN", ""),
        "control_key": row.get("BKONT", ""),
        "reference_details": row.get("BKREF", ""),
        "bank_country": row.get("BANKS", ""),
        "bank_key": row.get("BANKL", ""),
        "bank_name": "", "street": "", "city": "", "bank_branch": "", "swift_bic": "",
    }


def _partner_eq(a: str, b: str) -> bool:
    a, b = _norm_id(a), _norm_id(b)
    return bool(a) and (a == b or a.lstrip("0") == b.lstrip("0"))


def _row_score(row: dict, ext: Extraction) -> int:
    """How strongly a BUT0BK row matches the document's extracted identifiers."""
    f = ext.fields
    s = 0
    d_iban = _norm_id(f.get("iban"))
    r_iban = _norm_id(row.get("IBAN"))
    if d_iban and r_iban and d_iban == r_iban:
        s += 100
    d_acc = _norm_id(f.get("account_number"))
    r_acc = _norm_id(row.get("BANKN"))
    if d_acc and r_acc and d_acc.lstrip("0") == r_acc.lstrip("0"):
        s += 40
    if d_iban and r_acc and r_acc.lstrip("0") and d_iban.endswith(r_acc.lstrip("0")):
        s += 40
    d_key = _norm_id(f.get("routing_aba")) or _norm_id(f.get("routing_aba_wires"))
    if d_key and _norm_id(row.get("BANKL")) == d_key:
        s += 20
    return s


def select_row(rows: list[dict], ext: Extraction, sap_bp: str = "") \
        -> tuple[dict | None, list[Finding], list[str]]:
    """Choose the BUT0BK row to compare against + reverse-lookup findings.
    With sap_bp: the partner's best-matching bank-detail row (others noted).
    Without: reverse-lookup by the document's own identifiers.
    Returns (row|None, findings, other_partner_ids_for_notes)."""
    findings: list[Finding] = []
    if sap_bp:
        mine = [r for r in rows if _partner_eq(r.get("PARTNER", ""), sap_bp)]
        if not mine:
            findings.append(Finding("SAP-010", "WARNING", "WARNING",
                f"Business Partner {sap_bp} has no bank details in the export."))
            return None, findings, []
        mine.sort(key=lambda r: _row_score(r, ext), reverse=True)
        best = mine[0]
        if len(mine) > 1:
            findings.append(Finding("SAP-011", "NOTE", None,
                f"Business Partner {sap_bp} has {len(mine)} bank details on file; "
                f"compared against the best-matching one (Bank Details ID "
                f"{best.get('BKVID', '?')})."))
        return best, findings, []
    # reverse lookup: which partner(s) already hold this document's account?
    hits = [r for r in rows if _row_score(r, ext) >= 40]
    partners = sorted({r.get("PARTNER", "") for r in hits if r.get("PARTNER")})
    if not hits:
        findings.append(Finding("SAP-012", "NOTE", None,
            "The document's bank details were not found under any partner in the export."))
        return None, findings, []
    if len(partners) > 1:
        findings.append(Finding("SAP-013", "WARNING", "WARNING",
            f"These bank details already exist under {len(partners)} different "
            f"partners ({', '.join(partners[:6])}) — possible duplicate/shared account."))
    else:
        findings.append(Finding("SAP-012", "NOTE", None,
            f"These bank details already exist under partner {partners[0]}."))
    hits.sort(key=lambda r: _row_score(r, ext), reverse=True)
    return hits[0], findings, partners


# --- BUT000: BP general-data comparison (W-9 / tax docs) ------------------------
def _bp_names(row: dict) -> list[str]:
    parts = [row.get(k, "") for k in ("NAME_ORG1", "NAME_ORG2", "NAME_ORG3", "NAME_ORG4")]
    joined = " ".join(p for p in parts if p).strip()
    out = [n for n in (joined, row.get("BU_FULLNAME", ""),
                       (row.get("NAME_LAST", "") + " " + row.get("NAME_FIRST", "")).strip())
           if n]
    return out


def _is_individual(row: dict) -> bool | None:
    """SAP partner category: 1 = person, 2 = organization; Natural Person flag
    is the tiebreak. None when neither is present."""
    cat = _norm_id(row.get("TYPE"))
    if cat == "1" or _norm_id(row.get("NATPERS")) in ("x", "1"):
        return True
    if cat == "2":
        return False
    return None


def select_bp(rows: list[dict], sap_bp: str = "") -> dict | None:
    if sap_bp:
        for r in rows:
            if _partner_eq(r.get("PARTNER", ""), sap_bp):
                return r
        return None
    return rows[0] if len(rows) == 1 else None


def compare_bp(ext: Extraction, row: dict, policy: str = "masked") \
        -> tuple[list[Finding], list[dict]]:
    """BUT000 (BP general data) vs a W-9/W-8 document — name + entity type +
    central/archiving block. Rows share the {field,doc,sap,status,note} shape so
    the existing run.html table renders them unchanged."""
    f = ext.fields
    findings: list[Finding] = []
    rows: list[dict] = []

    def add(field, doc, sap, status, note=""):
        rows.append({"field": field, "doc": str(doc or ""), "sap": str(sap or ""),
                     "status": status, "note": note})

    # Line 1 (legal name) vs BP names — lenient (legal suffix / word order)
    d_name = str(f.get("line1_name") or f.get("account_holder") or "")
    sap_names = _bp_names(row)
    if d_name and sap_names:
        eq = any(names_materially_equal(d_name, s)
                 or _norm_name(d_name) in _norm_name(s)
                 or _norm_name(s) in _norm_name(d_name) for s in sap_names)
        add("Name (Line 1)", d_name, sap_names[0], "match" if eq else "MISMATCH")
        if not eq:
            findings.append(Finding("SAP-030", "CRITICAL", "NEED_MANUAL_REVIEW",
                f"W-9 name '{d_name}' does not match the SAP partner "
                f"'{sap_names[0]}' — do not process."))
    elif d_name or sap_names:
        add("Name (Line 1)", d_name, sap_names[0] if sap_names else "", "only-one-side")

    # Entity type: W-9 classification (individual vs entity) vs BP category
    cls = str(f.get("line3_classification") or "").lower()
    doc_individual = ("individual" in cls or "sole prop" in cls) if cls else None
    sap_individual = _is_individual(row)
    if doc_individual is not None and sap_individual is not None:
        eq = doc_individual == sap_individual
        add("Partner type",
            "individual" if doc_individual else "organization",
            "individual" if sap_individual else "organization",
            "match" if eq else "MISMATCH")
        if not eq:
            findings.append(Finding("SAP-031", "WARNING", "WARNING",
                "W-9 tax classification and the SAP partner category disagree "
                "(individual vs organization) — verify."))

    # Central / archiving block on a matched partner
    if _norm_id(row.get("XBLCK")).lower() in ("x", "1"):
        add("Central block", "", "X", "sap-only", "BP is centrally blocked")
        findings.append(Finding("SAP-032", "WARNING", "WARNING",
            f"SAP partner {row.get('PARTNER', '')} is centrally blocked."))
    if _norm_id(row.get("XDELE")).lower() in ("x", "1"):
        add("Archiving flag", "", "X", "sap-only", "BP flagged for archiving")
        findings.append(Finding("SAP-032", "WARNING", "WARNING",
            f"SAP partner {row.get('PARTNER', '')} is flagged for archiving/deletion."))

    add("Business Partner", "", row.get("PARTNER", ""), "sap-only")
    if not any(r["status"] == "MISMATCH" for r in rows) and rows:
        findings.append(Finding("SAP-000", "NOTE", None,
                                "SAP BP comparison: all comparable fields match."))
    return findings, rows


def register_secrets(ext: Extraction, sap_fields: dict) -> None:
    """Mirror sap_compare's vault registration for the SELECTED row only."""
    for k, kind in (("bank_account", "account_number"), ("iban", "iban")):
        v = sap_fields.get(k)
        if v:
            ext.vault.register(kind, v)
