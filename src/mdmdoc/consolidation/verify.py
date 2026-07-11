"""verify.py — post-write verification passes B/C/D.

Every consolidation re-proves itself on the SAVED file:

  B (cell check)   reopen the output with a fresh reader, resolve every
                   column from the header row with THIS module's own scan
                   (no shared code path with the writer) and assert each
                   planned cell holds exactly the planned value.
  C (integrity)    pre-existing content is untouched: per-sheet value hashes
                   over the pre-write row ranges must match the pre-write
                   snapshot, the new rows sit exactly in the planned range,
                   nothing exists below it, _COMMENT stays empty, and the
                   sheet inventory is unchanged.
  D (round-trip)   read the just-written vendor BACK out of the file by
                   SOURCE_ID and diff it against an independently re-derived
                   row-set (fresh parse of the source) — catches consistent
                   writer/plan bugs that B and C would both miss.

Any failed pass marks the result "blocked": the download endpoint refuses to
serve the file until a verified output exists.
"""
from __future__ import annotations

from pathlib import Path

from .template_io import cell_text, sheet_last_used, sheet_value_hash

# metadata columns with values a vendor row legitimately carries beyond the
# re-derived field set
_METADATA_TECHS = {"_ACTION_CODE", "SOURCE_ID", "SOURCE_ADDRNUMBER",
                   "ASSIGNMENT_ID", "PARTNER"}

# error messages quote cell values; TIN columns are masked there like in
# every other JSON artifact (the full value's only home is the workbook)
_TIN_TECHS = {"STCD1", "STCD2", "STCD3", "STCD4", "STCEG", "TAXNUM", "TAXNUMXL"}


def _quote(tech: str, value: str) -> str:
    if value and tech in _TIN_TECHS:
        from ..privacy import mask
        return repr(mask("tin", value))
    return repr(value)

# key column that distinguishes multi-row sheets of one vendor
_ROW_KEYS = {
    "LFB1 - Company Code (Supplier)": "BUKRS",
    "LFM1 - Purchasing Org Data": "EKORG",
    "DFKKBPTAXNUM - Tax Number": "TAXTYPE",
    "BUT0ID - Identifier": "TYPE",
}

# The vendor name may be SPILLED across a field family (NAME1..4 / NAME_ORG1..4)
# to fit SAP field lengths (see consolidation.fieldfit), and the split point
# differs per sheet because the field lengths differ. So compare the
# RECONSTRUCTED full name per sheet, not just the first field.
_NAME_FAMILIES = (
    ("LFA1 - Supplier General", ("NAME1", "NAME2", "NAME3", "NAME4")),
    ("BUT000 - General", ("NAME_ORG1", "NAME_ORG2", "NAME_ORG3", "NAME_ORG4")),
    ("ADRC - Address", ("NAME1", "NAME2", "NAME3", "NAME4")),
)


def _own_headers(ws, max_scan: int = 8):
    """Independent header scan (Pass B must not trust the writer's map)."""
    for r in range(1, max_scan + 1):
        row = ws[r]
        cells = {}
        anchored = False
        for c in row:
            t = cell_text(c.value)
            if not t:
                continue
            if t in ("_COMMENT", "SOURCE_ID"):
                anchored = True
            if t not in cells:
                cells[t] = c.column
        if anchored:
            return r, cells
    return None, {}


def _pass_cell_check(wb, write_plan: list[dict]) -> dict:
    errors = []
    headers: dict[str, tuple] = {}
    for c in write_plan:
        sheet = c["sheet"]
        if sheet not in headers:
            if sheet not in wb.sheetnames:
                errors.append(f"{sheet!r}: sheet missing from output")
                headers[sheet] = (None, {})
                continue
            headers[sheet] = _own_headers(wb[sheet])
        hrow, hmap = headers[sheet]
        if hrow is None:
            errors.append(f"{sheet!r}: no header row found in output")
            continue
        col = hmap.get(c["tech"])
        if col is None:
            errors.append(f"{sheet}!{c['tech']}: column not found in output header")
            continue
        if col != c["col"]:
            errors.append(
                f"{sheet}!{c['tech']}: writer used column {c['col']}, "
                f"independent header scan says {col} — column map drift"
            )
            continue
        actual = cell_text(wb[sheet].cell(row=c["row"], column=col).value)
        if actual != c["value"]:
            errors.append(
                f"{sheet}!R{c['row']} {c['tech']}: expected "
                f"{_quote(c['tech'], c['value'])}, file holds "
                f"{_quote(c['tech'], actual)}"
            )
    return {"pass": "cell_check", "checked": len(write_plan), "errors": errors}


def _pass_integrity(wb, pre_snapshot: dict, write_plan: list[dict]) -> dict:
    errors = []
    added_rows: dict[str, set] = {}
    for c in write_plan:
        added_rows.setdefault(c["sheet"], set()).add(c["row"])

    pre_names = list(pre_snapshot.keys())
    if list(wb.sheetnames) != pre_names:
        errors.append(
            f"sheet inventory changed: {len(pre_names)} sheets before, "
            f"{len(wb.sheetnames)} after"
        )

    for sheet, pre in pre_snapshot.items():
        if sheet not in wb.sheetnames:
            errors.append(f"{sheet!r}: sheet disappeared from output")
            continue
        ws = wb[sheet]
        pre_rows, pre_sha = pre["rows"], pre["sha"]
        if sheet_value_hash(ws, pre_rows) != pre_sha:
            errors.append(
                f"{sheet!r}: pre-existing rows 1..{pre_rows} were MODIFIED"
            )
        added = added_rows.get(sheet, set())
        expected_last = pre_rows + len(added) if added else pre_rows
        actual_last = sheet_last_used(ws)
        if added:
            lo, hi = min(added), max(added)
            if lo != pre_rows + 1 or hi != expected_last:
                errors.append(
                    f"{sheet!r}: new rows occupy {lo}..{hi}, planned "
                    f"{pre_rows + 1}..{expected_last}"
                )
        if actual_last != expected_last:
            errors.append(
                f"{sheet!r}: last used row is {actual_last}, expected "
                f"{expected_last} — stray or missing content"
            )
        if added:
            _, hmap = _own_headers(ws)
            ccol = hmap.get("_COMMENT")
            if ccol:
                for r in sorted(added):
                    if cell_text(ws.cell(row=r, column=ccol).value):
                        errors.append(
                            f"{sheet!r}!R{r}: _COMMENT is non-empty — SAP "
                            "would skip this row on upload"
                        )
    return {"pass": "integrity", "errors": errors}


def _vendor_rows_from_file(ws, hmap, hrow, source_id: str) -> list[dict]:
    sid_col = hmap.get("SOURCE_ID") or hmap.get("SOURCE_ADDRNUMBER")
    rows = []
    for r in range(hrow + 1, sheet_last_used(ws) + 1):
        if sid_col is None:
            continue
        if cell_text(ws.cell(row=r, column=sid_col).value) != source_id:
            continue
        fields = {}
        for tech, col in hmap.items():
            v = cell_text(ws.cell(row=r, column=col).value)
            if v:
                fields[tech] = v
        rows.append(fields)
    return rows


def _pass_roundtrip(wb, fresh_rows_by_sheet: dict, source_id: str) -> dict:
    """Diff independently re-derived rows against what the file holds for
    this SOURCE_ID."""
    errors = []
    notes = []
    if not source_id:
        return {"pass": "roundtrip", "notes": [],
                "errors": ["no SOURCE_ID to select rows by"]}

    file_rows_by_sheet: dict[str, list[dict]] = {}
    for sheet, fresh_rows in fresh_rows_by_sheet.items():
        if sheet not in wb.sheetnames:
            errors.append(f"{sheet!r}: sheet missing from output")
            continue
        ws = wb[sheet]
        hrow, hmap = _own_headers(ws)
        if hrow is None:
            errors.append(f"{sheet!r}: no header row in output")
            continue
        if not (hmap.get("SOURCE_ID") or hmap.get("SOURCE_ADDRNUMBER")):
            # cannot select this vendor's rows without an id column; pass B
            # already verified every planned cell there — note, don't block
            notes.append(f"{sheet!r}: no SOURCE_ID column — round-trip "
                         "skipped for this sheet (cell check covers it)")
            continue
        file_rows = _vendor_rows_from_file(ws, hmap, hrow, source_id)
        file_rows_by_sheet[sheet] = file_rows

        expected = []
        for row in fresh_rows:
            exp = {}
            for tech, value in row.items():
                s = cell_text(value)
                if s:
                    exp[tech] = s
            expected.append(exp)

        if len(file_rows) < len(expected):
            errors.append(
                f"{sheet!r}: file has {len(file_rows)} row(s) for "
                f"SOURCE_ID {source_id!r}, re-derivation expects {len(expected)}"
            )
            continue

        key = _ROW_KEYS.get(sheet)
        for exp in expected:
            if key and exp.get(key):
                match = [fr for fr in file_rows if fr.get(key) == exp[key]]
                if not match:
                    errors.append(
                        f"{sheet!r}: no row with {key}={exp[key]!r} for this vendor"
                    )
                    continue
                fr = match[0]
            else:
                fr = file_rows[expected.index(exp)] if len(file_rows) >= len(expected) else None
                if fr is None:
                    continue
            for tech, val in exp.items():
                got = fr.get(tech, "")
                if got != val:
                    pos = _first_diff(val, got)
                    errors.append(
                        f"{sheet}.{tech}: re-derived {_quote(tech, val)} vs "
                        f"file {_quote(tech, got)} (first diff at char {pos})"
                    )
            unexpected = set(fr) - set(exp) - _METADATA_TECHS
            for tech in sorted(unexpected):
                errors.append(
                    f"{sheet}.{tech}: file holds {_quote(tech, fr[tech])} but "
                    "the re-derivation planned nothing for this column"
                )

    # cross-sheet invariants
    names = {}
    for sheet, fields in _NAME_FAMILIES:
        rows = file_rows_by_sheet.get(sheet) or []
        if not rows:
            continue
        full = " ".join(" ".join(cell_text(rows[0].get(f)).split())
                        for f in fields).strip()
        if full:
            names[sheet] = full
    if len(set(names.values())) > 1:
        errors.append(f"vendor name differs across sheets: {names}")

    lfb1 = file_rows_by_sheet.get("LFB1 - Company Code (Supplier)") or []
    if len(lfb1) > 1:
        broadcast_keys = set().union(*(set(r) for r in lfb1)) - {"BUKRS"}
        for k in sorted(broadcast_keys):
            vals = {r.get(k, "") for r in lfb1}
            if len(vals) > 1:
                errors.append(
                    f"LFB1 broadcast {k} differs across company-code rows: "
                    f"{sorted(vals)}"
                )

    partners = set()
    for rows in file_rows_by_sheet.values():
        for r in rows:
            if "PARTNER" in r:
                partners.add(r["PARTNER"])
    if len(partners) > 1:
        errors.append(f"PARTNER differs across rows: {sorted(partners)}")

    return {"pass": "roundtrip", "errors": errors, "notes": notes}


def _first_diff(a: str, b: str) -> int:
    for i, (x, y) in enumerate(zip(a, b)):
        if x != y:
            return i + 1
    return min(len(a), len(b)) + 1


def verify_output(out_path: Path, write_plan: list[dict], pre_snapshot: dict,
                  fresh_by_vendor: dict | None = None) -> dict:
    """Run passes B, C and (per vendor) D on the saved file.

    fresh_by_vendor: {source_id: rows_by_sheet} — one independently
    re-derived row-set per vendor written in this batch."""
    import openpyxl
    wb = openpyxl.load_workbook(out_path)
    try:
        passes = [
            _pass_cell_check(wb, write_plan),
            _pass_integrity(wb, pre_snapshot, write_plan),
        ]
        for sid, fresh_rows in (fresh_by_vendor or {}).items():
            p = _pass_roundtrip(wb, fresh_rows, sid)
            p["source_id"] = sid
            passes.append(p)
    finally:
        wb.close()
    blocked = any(p["errors"] for p in passes)
    return {
        "schema": "mdmdoc.consolidate.v1",
        "status": "blocked" if blocked else "verified",
        "passes": passes,
    }
