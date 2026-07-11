"""template_io.py — load, validate, and APPEND to a Business Partner
consolidation template (BusinessPartnerTemplate.xlsx).

The operator uploads their own copy of the template and it may already carry
vendors. Everything here is append-only by construction:

- rows are written strictly AFTER the last used row of each sheet (a row is
  "used" when any cell holds a value; formatting-only ghost rows don't count);
- the uploaded file itself is never modified — save_to() refuses to write
  over the input path;
- the _COMMENT column is never written (the template's Read Me: any content
  in _COMMENT makes SAP skip the entire row on upload);
- _ACTION_CODE and SOURCE_ID metadata land only where the sheet actually has
  those columns (the real template's BUT000 sheet has no _ACTION_CODE).

Header rows are DETECTED, not assumed: the row within the first
MAX_HEADER_SCAN rows that carries a `_COMMENT` or `SOURCE_ID` cell is the
technical header (row 2 in the SAP-delivered template; row 6 in the
SUPPLIER.xlsx example layout). Anything else fails loud with BPTemplateError
before a single cell is written.
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import zipfile
from pathlib import Path

MAX_HEADER_SCAN = 8
_HEADER_ANCHORS = ("_COMMENT", "SOURCE_ID")

# openpyxl saves string cells as INLINE strings (<is><t>…). SAP's ABAP xlsx
# reader (the MDG/S4 "Create Data Import") only reads SHARED strings (t="s"
# referencing xl/sharedStrings.xml) — Excel's default — so an openpyxl-saved
# workbook fails to import with a raw CX_ROOT. This converts every inline
# string back to a shared string post-save. The inner content model of <is>
# and <si> is identical, so the inner bytes move verbatim (styles/xml:space/
# rich runs preserved); only t="inlineStr" -> t="s"><v>idx</v>.
# non-empty inline string: <c … t="inlineStr"><is>INNER</is></c> (INNER moves
# verbatim into a shared <si>). `.+?` so an EMPTY <is></is> is left to the
# empty pattern below rather than becoming an empty shared string.
_INLINE_STR = re.compile(
    rb'<c([^>]*?)\st="inlineStr"([^>]*?)>\s*<is>(.+?)</is>\s*</c>', re.S)
# empty string cell — openpyxl emits it as EITHER a self-closing
# <c … t="inlineStr"/> OR an open/close <c … t="inlineStr"></c> (optionally an
# empty <is/>); depends on the openpyxl version. Both become a blank cell.
_INLINE_EMPTY = re.compile(
    rb'<c([^>]*?)\st="inlineStr"([^>]*?)(?:\s*/>|>\s*(?:<is\s*/>|<is>\s*</is>)?\s*</c>)',
    re.S)


def _inline_to_shared(path: Path) -> None:
    with zipfile.ZipFile(path) as z:
        parts = {i.filename: z.read(i.filename) for i in z.infolist()}
    order: list[bytes] = []
    index: dict[bytes, int] = {}
    total = [0]

    def repl(m):
        inner = m.group(3)
        if inner not in index:
            index[inner] = len(order)
            order.append(inner)
        total[0] += 1
        return (b'<c' + m.group(1) + m.group(2) + b' t="s"><v>'
                + str(index[inner]).encode() + b'</v></c>')

    changed = False
    for name in list(parts):
        if name.startswith("xl/worksheets/") and name.endswith(".xml"):
            new = _INLINE_STR.sub(repl, parts[name])
            new = _INLINE_EMPTY.sub(rb'<c\1\2/>', new)   # empty string -> blank cell
            if new != parts[name]:
                parts[name] = new
                changed = True
    if not changed:
        return
    if not order:
        # only empty inline cells were present (no shared table needed), but the
        # sheet XML changed — rewrite the zip so the blanks land
        _rewrite_zip(path, parts)
        return
    parts["xl/sharedStrings.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'count="' + str(total[0]).encode() + b'" uniqueCount="'
        + str(len(order)).encode() + b'">'
        + b''.join(b'<si>' + s + b'</si>' for s in order) + b'</sst>')
    ct_override = (
        b'<Override PartName="/xl/sharedStrings.xml" ContentType='
        b'"application/vnd.openxmlformats-officedocument.spreadsheetml.'
        b'sharedStrings+xml"/>')
    ct = parts["[Content_Types].xml"]
    if b"sharedStrings.xml" not in ct:
        parts["[Content_Types].xml"] = ct.replace(
            b"</Types>", ct_override + b"</Types>")
    rels = parts["xl/_rels/workbook.xml.rels"]
    if b"sharedStrings.xml" not in rels:
        ids = [int(x) for x in re.findall(rb'Id="rId(\d+)"', rels)] or [0]
        newid = b"rId" + str(max(ids) + 1).encode()
        rel = (b'<Relationship Id="' + newid + b'" Type="http://schemas.'
               b'openxmlformats.org/officeDocument/2006/relationships/sharedStrings"'
               b' Target="sharedStrings.xml"/>')
        parts["xl/_rels/workbook.xml.rels"] = rels.replace(
            b"</Relationships>", rel + b"</Relationships>")
    _rewrite_zip(path, parts)


def _rewrite_zip(path: Path, parts: dict) -> None:
    tmp = str(path) + ".sstmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in parts.items():
            z.writestr(name, data)
    os.replace(tmp, path)

# Sheets a vendor row-set targets; validate() requires them to exist with a
# detectable header row carrying these anchor columns.
REQUIRED_SHEETS = {
    "LFA1 - Supplier General": ("LIFNR", "NAME1"),
    "BUT000 - General": ("PARTNER", "NAME_ORG1"),
    "BUT100 - Role": ("RLTYP",),
    "ADRC - Address": ("NAME1",),
    "BUT0BK - Bank Account": ("BANKS", "BANKL", "BANKN"),
    "LFB1 - Company Code (Supplier)": ("BUKRS",),
    "LFM1 - Purchasing Org Data": ("EKORG",),
}

# Name columns used for duplicate-vendor detection.
_NAME_COLUMNS = (
    ("LFA1 - Supplier General", "NAME1"),
    ("BUT000 - General", "NAME_ORG1"),
)

# Customer tables (KN*) — hidden in the output so they don't clutter a vendor
# consolidation. SAP ignores hidden sheets on import (per the template Read Me).
def _is_customer_sheet(name: str) -> bool:
    return name.split(" - ")[0].strip().upper().startswith("KN")


class BPTemplateError(ValueError):
    """The uploaded workbook is not a usable BP consolidation template."""


def cell_text(v) -> str:
    """Stringify a cell value the way SAP raw format expects: no trailing
    '.0' on integer floats, no True/False literals, whitespace stripped."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "X" if v else ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def normalized_name(v) -> str:
    return " ".join(cell_text(v).upper().split())


# Business address / language columns that must STAY VISIBLE even when empty —
# operators fill them and SAP requires several. Hiding was only ever meant for
# customer sheets + SAP-internal empties, not these.
_NEVER_HIDE = {
    "SPRAS", "LANGU", "BU_LANGU", "LANGU_CORR",
    "COUNTRY", "LAND1", "REGION", "REGIO",
    "POST_CODE1", "PSTLZ", "CITY1", "CITY2", "ORT01", "ORT02",
    "STREET", "STRAS", "STR_SUPPL1", "STR_SUPPL2", "STR_SUPPL3",
}


def _hide_empty_columns(wb, key_fields_by_sheet: dict) -> None:
    """Hide every NON-KEY, NON-whitelisted column that is empty across all data
    rows.

    SAP validates every VISIBLE column of a written row and ignores hidden ones
    (template Read Me). An empty visible non-key column triggers either a hard
    error (empty timestamp UPTIM / GUID BP_BANK_GUID — SAP can't default it) or
    a warning (empty date → "Assumed value 00.00.0000"). Hiding empty non-key
    columns removes both. KEY columns (underlined) and business address/language
    columns (`_NEVER_HIDE`) are NEVER hidden even when blank — operators fill
    them and SAP requires several; the keys we DO fill (SOURCE_ID, ASSIGNMENT_ID,
    DATE_FROM, CONSNUMBER, BKVID) plus any populated column stay visible too.
    """
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets:
        if ws.sheet_state == "hidden" or str(ws.title).startswith("!"):
            continue
        if ws.max_row < 3:                        # header only, no data rows
            continue
        keys = key_fields_by_sheet.get(ws.title) or set()
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=col).value
            header = str(header).strip() if header is not None else ""
            if not header or header in keys or header in _NEVER_HIDE:
                continue                          # keep unlabeled / KEY / business columns
            used = any(cell_text(ws.cell(row=r, column=col).value)
                       for r in range(3, ws.max_row + 1))
            if not used:
                ws.column_dimensions[get_column_letter(col)].hidden = True


def sheet_last_used(ws, from_row: int = 1) -> int:
    """Last row >= from_row with any value; from_row-1 when none."""
    last = from_row - 1
    for r, row in enumerate(ws.iter_rows(min_row=from_row, values_only=True),
                            start=from_row):
        if any(cell_text(v) for v in row):
            last = r
    return last


def sheet_value_hash(ws, upto_row: int) -> str:
    """SHA-256 over the values of rows 1..upto_row (canonical JSON per row).
    Shared by the pre-write snapshot and the post-write integrity check so
    the two sides can never drift on serialization details."""
    acc = hashlib.sha256()
    if upto_row < 1:
        return acc.hexdigest()
    for row in ws.iter_rows(min_row=1, max_row=upto_row, values_only=True):
        texts = [cell_text(v) for v in row]
        while texts and texts[-1] == "":
            texts.pop()
        acc.update(json.dumps(texts, ensure_ascii=False).encode("utf-8"))
        acc.update(b"\n")
    return acc.hexdigest()


def snapshot_workbook(wb) -> dict:
    """{sheet: {"rows": last_used_row, "sha": hex}} across EVERY sheet —
    the integrity baseline proving pre-existing content survived untouched."""
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        last = sheet_last_used(ws)
        out[name] = {"rows": last, "sha": sheet_value_hash(ws, last)}
    return out


def find_header_row(ws, max_scan: int = MAX_HEADER_SCAN):
    """(header_row, {tech: col}) or (None, {}) — the first row in the scan
    window carrying a _COMMENT or SOURCE_ID cell is the technical header."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), start=1):
        texts = [cell_text(c.value) for c in row]
        if any(t in _HEADER_ANCHORS for t in texts):
            idx = {}
            for c in row:
                t = cell_text(c.value)
                if t and t not in idx:
                    idx[t] = c.column
            return r, idx
    return None, {}


class BPTemplate:
    """An open BP template workbook with append-only write access."""

    def __init__(self, path):
        self.path = Path(path)
        if not self.path.exists():
            raise BPTemplateError(f"template not found: {self.path}")
        import openpyxl
        try:
            self.wb = openpyxl.load_workbook(self.path)
        except Exception as exc:
            raise BPTemplateError(
                f"cannot open workbook ({exc.__class__.__name__}: {exc})"
            ) from exc
        self._headers: dict[str, dict] = {}
        self._header_rows: dict[str, int] = {}
        self._key_fields: dict[str, set] | None = None

    # --- structure -----------------------------------------------------------

    def has_sheet(self, sheet: str) -> bool:
        return sheet in self.wb.sheetnames

    def key_fields(self) -> dict[str, set]:
        """{sheet: {tech, …}} — the KEY fields SAP requires, read from the
        template's own marking: key headers are UNDERLINED (bold+underline,
        fill FFDBD5BF). This is exactly the "подчёркнутые" the operator sees.
        Cached."""
        if self._key_fields is not None:
            return self._key_fields
        out: dict[str, set] = {}
        for sheet in self.wb.sheetnames:
            try:
                self._index(sheet)
            except BPTemplateError:
                continue
            hrow = self._header_rows[sheet]
            ws = self.wb[sheet]
            keys = set()
            for c in ws[hrow]:
                if c.value and c.font and c.font.underline:
                    keys.add(str(c.value).strip())
            out[sheet] = keys
        self._key_fields = out
        return out

    def _index(self, sheet: str) -> None:
        if sheet in self._headers:
            return
        if sheet not in self.wb.sheetnames:
            raise BPTemplateError(f"sheet missing from template: {sheet!r}")
        row, idx = find_header_row(self.wb[sheet])
        if row is None:
            raise BPTemplateError(
                f"{sheet!r}: no technical header row (_COMMENT/SOURCE_ID) in "
                f"the first {MAX_HEADER_SCAN} rows — not a BP consolidation template"
            )
        self._header_rows[sheet] = row
        self._headers[sheet] = idx

    def header_row(self, sheet: str) -> int:
        self._index(sheet)
        return self._header_rows[sheet]

    def column_for(self, sheet: str, tech: str):
        """build_rows() column protocol: int column or None."""
        self._index(sheet)
        return self._headers[sheet].get(tech)

    def validate(self) -> None:
        problems = []
        for sheet, anchors in REQUIRED_SHEETS.items():
            try:
                self._index(sheet)
            except BPTemplateError as exc:
                problems.append(str(exc))
                continue
            for tech in anchors:
                if tech not in self._headers[sheet]:
                    problems.append(
                        f"{sheet!r}: header row {self._header_rows[sheet]} "
                        f"has no {tech!r} column"
                    )
        if problems:
            raise BPTemplateError(
                "not a usable Business Partner template: " + "; ".join(problems)
            )

    # --- content -------------------------------------------------------------

    def last_used_row(self, sheet: str) -> int:
        """Last row holding any value (headers count), so appends always land
        after BOTH the header block and every pre-existing data row."""
        self._index(sheet)
        ws = self.wb[sheet]
        return max(self._header_rows[sheet], sheet_last_used(ws))

    def data_rows_used(self, sheet: str) -> int:
        return max(0, self.last_used_row(sheet) - self.header_row(sheet))

    def clear_data_rows(self) -> dict:
        """Wipe every existing vendor row (below the technical header) on the
        real data sheets, giving a guaranteed-clean template to append into.

        Clears by NULLING cell values, not delete_rows: openpyxl's delete_rows
        corrupts styles/merged cells and shifts the sheet inventory that verify
        pass C guards; value-nulling keeps structure and yields the same clean
        append point (sheet_last_used/value-hash are value-based). Only sheets
        whose header carries SOURCE_ID are touched — this excludes '! Read Me !'
        (its only anchor is _COMMENT). The uploaded file is never mutated (this
        edits the in-memory workbook; the caller saves a cleaned copy)."""
        cleared_rows = 0
        sheets_touched = []
        for sheet in list(self.wb.sheetnames):
            try:
                self._index(sheet)
            except BPTemplateError:
                continue
            if "SOURCE_ID" not in self._headers[sheet]:
                continue
            ws = self.wb[sheet]
            header = self._header_rows[sheet]
            last = sheet_last_used(ws)
            n = 0
            for r in range(header + 1, last + 1):
                row_has_value = any(
                    cell_text(ws.cell(row=r, column=c).value)
                    for c in range(1, ws.max_column + 1))
                if not row_has_value:
                    continue
                for c in range(1, ws.max_column + 1):
                    if ws.cell(row=r, column=c).value is not None:
                        ws.cell(row=r, column=c).value = None
                n += 1
            if n:
                cleared_rows += n
                sheets_touched.append(sheet)
        return {"rows": cleared_rows, "sheets": len(sheets_touched)}

    def _sheets_with_column(self, tech: str):
        for sheet in self.wb.sheetnames:
            try:
                self._index(sheet)
            except BPTemplateError:
                continue
            col = self._headers[sheet].get(tech)
            if col:
                yield sheet, col

    def existing_source_ids(self) -> set[str]:
        out = set()
        for sheet, col in self._sheets_with_column("SOURCE_ID"):
            ws = self.wb[sheet]
            first = self._header_rows[sheet] + 1
            for row in ws.iter_rows(min_row=first, min_col=col, max_col=col,
                                    values_only=True):
                v = cell_text(row[0])
                if v:
                    out.add(v)
        return out

    def existing_vendor_names(self) -> set[str]:
        out = set()
        for sheet, tech in _NAME_COLUMNS:
            try:
                self._index(sheet)
            except BPTemplateError:
                continue
            col = self._headers[sheet].get(tech)
            if not col:
                continue
            ws = self.wb[sheet]
            first = self._header_rows[sheet] + 1
            for row in ws.iter_rows(min_row=first, min_col=col, max_col=col,
                                    values_only=True):
                v = normalized_name(row[0])
                if v:
                    out.add(v)
        return out

    def snapshot(self) -> dict:
        return snapshot_workbook(self.wb)

    # --- writing -------------------------------------------------------------

    def plan_rows(self, rows_by_sheet: dict) -> list[dict]:
        """Dry-run of append_rows: the exact cells a write would touch,
        without touching them. Raises BPTemplateError on anything unsafe."""
        return self._place(rows_by_sheet, write=False)

    def append_rows(self, rows_by_sheet: dict) -> list[dict]:
        """Append one vendor's rows after existing data. Returns the write
        plan actually executed: [{sheet,row,col,tech,value}, ...]."""
        # A failed pre-flight must leave the workbook untouched.
        self._place(rows_by_sheet, write=False)
        return self._place(rows_by_sheet, write=True)

    def _place(self, rows_by_sheet: dict, write: bool) -> list[dict]:
        plan = []
        for sheet, rows in rows_by_sheet.items():
            self._index(sheet)
            ws = self.wb[sheet]
            start = self.last_used_row(sheet) + 1
            for i, fields in enumerate(rows):
                r = start + i
                for tech, value in fields.items():
                    if value is None:
                        continue
                    if tech == "_COMMENT":
                        raise BPTemplateError(
                            "_COMMENT must never be written — SAP skips rows "
                            "with _COMMENT content on upload"
                        )
                    col = self._headers[sheet].get(tech)
                    if col is None:
                        raise BPTemplateError(
                            f"{sheet!r} has no column {tech!r} — refusing to "
                            "guess a target cell"
                        )
                    s = cell_text(value)
                    if not s:
                        continue
                    if s.startswith("="):
                        # openpyxl types '='-prefixed strings as FORMULAS: the
                        # verify passes would compare the formula text and pass
                        # while Excel/SAP evaluates something else entirely
                        # (also a formula-injection vector from form values).
                        raise BPTemplateError(
                            f"{sheet}.{tech}: value {s!r} starts with '=' — "
                            "Excel would treat it as a formula; refusing"
                        )
                    if write:
                        ws.cell(row=r, column=col, value=s)
                    plan.append({"sheet": sheet, "row": r, "col": col,
                                 "tech": tech, "value": s})
        return plan

    def save_to(self, out_path, hide_empty: bool = True) -> Path:
        out_path = Path(out_path)
        if out_path.exists() and out_path.samefile(self.path):
            raise BPTemplateError("refusing to overwrite the uploaded template")
        if out_path.resolve() == self.path.resolve():
            raise BPTemplateError("refusing to overwrite the uploaded template")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        # hide customer (KN*) sheets so a vendor consolidation isn't cluttered;
        # SAP ignores hidden sheets on import
        for ws in self.wb.worksheets:
            if _is_customer_sheet(ws.title):
                ws.sheet_state = "hidden"
        # hide empty non-key columns so SAP doesn't reject empty timestamp/GUID
        # cells or warn on empty dates (recomputed each save → re-uploads that
        # populate a previously-empty column make it visible again)
        if hide_empty:
            _hide_empty_columns(self.wb, self.key_fields())
        self.wb.save(out_path)
        # SAP's importer needs shared strings, not openpyxl's inline strings
        _inline_to_shared(out_path)
        return out_path

    def close(self) -> None:
        self.wb.close()
