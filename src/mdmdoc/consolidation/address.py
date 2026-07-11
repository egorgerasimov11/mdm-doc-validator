"""address.py — consolidation address completeness for SAP import.

Three post-build steps (run in _rows_for_vendor after apply_constants, before
fieldfit), all deterministic so the Pass-D round-trip verification still matches:

1. `add_international_version` — the form may carry a second "International
   Address" block in the local script (e.g. Chinese). Emit it as a 2nd ADRC row
   with the SAP address version NATION (C=Chinese, H=Hangul, K=Kanji).
2. `assign_source_addrnumber` — SOURCE_ADDRNUMBER = "1" on every address row.
   SAP rejects an INITIAL SOURCE_ADDRNUMBER for internal-numbering vendors
   ("Initial value for field ADR2-SOURCE_ADDRNUMBER not allowed"); it is the
   per-vendor address number. (SAP's SUPPLIER example leaves it blank, but that
   uses EXTERNAL numbering — it misled ЭТАП-5 into blanking it.)
3. `fill_region` — resolve REGIO (LFA1) / REGION (ADRC) from country+city+postal
   via the address-validator (offline, in-process). SAP requires REGION for many
   countries ("Enter a value for field REGION"); the converter can't derive it.
"""
from __future__ import annotations

from pathlib import Path

from .template_io import BPTemplateError, cell_text

_ADDR_SHEETS = ("ADRC - Address", "ADR2 - Phone", "ADR6 - E-Mail")
SOURCE_ADDRNUMBER_DEFAULT = "1"

# SAP international address version (ADRC.NATION) by country ISO2. Blank NATION is
# the default/romanized version; these carry the local-script version.
_NATION_BY_COUNTRY = {"CN": "C", "KR": "H", "JP": "K"}

# name / street / city fields replaced by the local-script version (the rest of
# the cloned ADRC row — SOURCE_ID, SOURCE_ADDRNUMBER, DATE_FROM, CONSNUMBER,
# COUNTRY, REGION, POST_CODE1, phone… — is inherited).
_INTL_CLEARED = ("NAME1", "NAME2", "NAME3", "NAME4", "STREET",
                 "STR_SUPPL1", "STR_SUPPL2", "STR_SUPPL3", "CITY1", "CITY2", "SORT2")


# --- 1. international (local-script) address version ------------------------

def _read_international_block(form_path) -> dict | None:
    """Read the form's 'International Address' block (local-script Name/Street/
    City). Label-anchored: find the 'International Address' header, then read the
    labelled fields from the value cell immediately to the right of each label."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(form_path, data_only=True, keep_vba=False)
    except Exception:
        return None
    try:
        if "2. Vendor Details" not in wb.sheetnames:
            return None
        ws = wb["2. Vendor Details"]
        header_row = header_col = None
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and \
                        c.value.strip().lower() == "international address":
                    header_row, header_col = c.row, c.column
                    break
            if header_row:
                break
        if not header_row:
            return None
        want = {"name 1": "name",
                "building number/street": "street", "building/street": "street",
                "street/house number": "street", "city": "city"}
        out: dict = {}
        for r in range(header_row + 1, header_row + 30):
            label = ws.cell(row=r, column=header_col).value
            if not isinstance(label, str):
                continue
            key = want.get(label.strip().lower())
            if key and key not in out:
                out[key] = cell_text(ws.cell(row=r, column=header_col + 1).value)
        return out or None
    finally:
        wb.close()


def add_international_version(rows_by_sheet: dict, form_path) -> dict:
    """Append a 2nd ADRC row for the local-script address version (NATION set by
    country). No-op when there is no ADRC row, no International Address block, or
    the country has no local-script version."""
    adrc_rows = rows_by_sheet.get("ADRC - Address")
    if not adrc_rows or not form_path:
        return rows_by_sheet
    base = adrc_rows[0]
    nation = _NATION_BY_COUNTRY.get(cell_text(base.get("COUNTRY")).upper())
    if not nation:
        return rows_by_sheet
    intl = _read_international_block(Path(form_path))
    if not intl or not cell_text(intl.get("name")):
        return rows_by_sheet
    row = dict(base)                      # inherit metadata + country/region/postal
    for f in _INTL_CLEARED:
        row.pop(f, None)                  # drop the romanized name/street/city
    row["NAME1"] = intl["name"]
    if intl.get("street"):
        row["STREET"] = intl["street"]
    if intl.get("city"):
        row["CITY1"] = intl["city"]
    row["NATION"] = nation
    adrc_rows.append(row)                 # AFTER the romanized row (row[0])
    return rows_by_sheet


# --- 2. address number -----------------------------------------------------

def assign_source_addrnumber(rows_by_sheet: dict) -> dict:
    """SOURCE_ADDRNUMBER = "1" on every address row — non-initial (SAP requires
    it), ≤10 chars, identical across ADRC/ADR2/ADR6 so phone/email bind to the
    address (and both ADRC version rows share the one address)."""
    for sheet in _ADDR_SHEETS:
        for row in rows_by_sheet.get(sheet, []):
            row["SOURCE_ADDRNUMBER"] = SOURCE_ADDRNUMBER_DEFAULT
    return rows_by_sheet


# --- 3. region via the address-validator -----------------------------------

_REGION_CTX = None
_REGION_CTX_TRIED = False


def _region_ctx():
    """Lazy, offline, in-process address-validator context (singleton)."""
    global _REGION_CTX, _REGION_CTX_TRIED
    if _REGION_CTX_TRIED:
        return _REGION_CTX
    _REGION_CTX_TRIED = True
    try:
        from addrval.engine.pipeline import EngineContext
        _REGION_CTX = EngineContext.load(online=False, websearch=False)
    except Exception:
        _REGION_CTX = None
    return _REGION_CTX


def _resolve_region(country_iso: str, city, postal, warnings_out=None):
    """Resolve the SAP region TWO independent ways and cross-check (accuracy):
    by CITY (city-admin1/medium) and by POSTAL (postcode/high). Trust an
    agreement; on disagreement prefer the postcode/high code + warn; otherwise
    whichever resolved."""
    ctx = _region_ctx()
    if ctx is None or not country_iso:
        return None
    city, postal = cell_text(city), cell_text(postal)
    if not (city or postal):
        return None
    try:
        from addrval.engine.sap_regions import resolve_region
        by_city = resolve_region(country_iso, city, "", "", ctx.sap, ctx.geo) if city else None
        by_postal = resolve_region(country_iso, "", "", postal, ctx.sap, ctx.geo) if postal else None
    except Exception:
        return None
    cc = by_city[0] if by_city else None
    cp = by_postal[0] if by_postal else None
    if cc and cp and cc != cp:
        if warnings_out is not None:
            warnings_out.append(
                f"REGION cross-check disagreed (city={cc}, postal={cp}); "
                f"used postal {cp} — verify")
        return cp                                     # postcode is the high-confidence source
    return cp or cc                                   # agreement, or whichever resolved


def fill_region(rows_by_sheet: dict, warnings_out: list | None = None) -> dict:
    """Fill LFA1.REGIO + ADRC.REGION from country+city+postal when the form gave
    no region. Pass the ROMANIZED city (a CJK city can't be matched)."""
    lfa1_rows = rows_by_sheet.get("LFA1 - Supplier General") or []
    adrc_rows = rows_by_sheet.get("ADRC - Address") or []
    lfa1 = lfa1_rows[0] if lfa1_rows else {}
    adrc = adrc_rows[0] if adrc_rows else {}
    if cell_text(lfa1.get("REGIO")) or cell_text(adrc.get("REGION")):
        return rows_by_sheet                          # form already supplied it
    country = cell_text(lfa1.get("LAND1")) or cell_text(adrc.get("COUNTRY"))
    city = adrc.get("CITY1") or lfa1.get("ORT01")
    postal = adrc.get("POST_CODE1") or lfa1.get("PSTLZ")
    code = _resolve_region(country, city, postal, warnings_out)
    if not code:
        if warnings_out is not None and country:
            warnings_out.append(
                f"REGION could not be resolved for {country} "
                f"{cell_text(city)} {cell_text(postal)} — fill it manually")
        return rows_by_sheet
    for r in lfa1_rows:
        if not cell_text(r.get("REGIO")):
            r["REGIO"] = code
    for r in adrc_rows:
        if not cell_text(r.get("REGION")):
            r["REGION"] = code
    return rows_by_sheet


# --- 4. phone country + district -------------------------------------------

def _has(columns, sheet, field):
    if columns is None:
        return True
    try:
        return columns.column_for(sheet, field) is not None
    except BPTemplateError:                        # sheet absent from this template
        return False


def fill_phone_country(rows_by_sheet: dict, columns=None) -> dict:
    """SAP validates a phone's area code against a country; fill ADR2/ADR3 COUNTRY
    from the vendor country ("Specify a valid country/area code")."""
    lfa1 = rows_by_sheet.get("LFA1 - Supplier General") or [{}]
    adrc = rows_by_sheet.get("ADRC - Address") or [{}]
    country = cell_text(lfa1[0].get("LAND1")) or cell_text(adrc[0].get("COUNTRY"))
    if not country:
        return rows_by_sheet
    for sheet in ("ADR2 - Phone", "ADR3 - Fax"):
        if not _has(columns, sheet, "COUNTRY"):
            continue
        for row in rows_by_sheet.get(sheet, []):
            if not cell_text(row.get("COUNTRY")):
                row["COUNTRY"] = country
    return rows_by_sheet


def _read_region_string(form_path) -> str:
    """Read the form's 'State/Province/Region' free-text (label-anchored)."""
    try:
        from sap_vendor_autoload.reader import SourceForm, DETAILS
        with SourceForm(Path(form_path)) as src:
            return cell_text(src.get_str(DETAILS, "E24"))
    except Exception:
        return ""


def fill_district(rows_by_sheet: dict, form_path, columns=None) -> dict:
    """Parse the district out of the form's region string (e.g.
    "CHAOYANG DISTRICT,BEIJING,CHINA" → "CHAOYANG DISTRICT") into ADRC.CITY2 +
    LFA1.ORT02 when empty. The address-validator has no district data."""
    adrc_rows = rows_by_sheet.get("ADRC - Address") or []
    if not adrc_rows or not form_path:
        return rows_by_sheet
    base = adrc_rows[0]
    if any(cell_text(r.get("CITY2")) for r in adrc_rows):
        return rows_by_sheet                          # already have a district
    reg = _read_region_string(form_path)
    if not reg:
        return rows_by_sheet
    import re
    parts = [p.strip() for p in re.split(r"[,;\n]", reg) if p.strip()]
    if len(parts) < 2:
        return rows_by_sheet                          # a single clean value IS the region, not a district
    try:
        from sap_vendor_autoload.transforms import iso_country
    except Exception:
        iso_country = lambda s: None
    city = cell_text(base.get("CITY1")).upper()
    country = cell_text(base.get("COUNTRY")).upper()
    region = cell_text(base.get("REGION")).upper()
    district = None
    for part in parts:
        pu = part.upper()
        if pu in (city, country, region) or (iso_country(part) or "").upper() == country:
            continue                                  # this segment is the city / country / region
        district = part
        break
    if not district:
        return rows_by_sheet
    if _has(columns, "ADRC - Address", "CITY2"):
        for r in adrc_rows:
            if not cell_text(r.get("CITY2")):
                r["CITY2"] = district
    if _has(columns, "LFA1 - Supplier General", "ORT02"):
        for r in rows_by_sheet.get("LFA1 - Supplier General", []):
            if not cell_text(r.get("ORT02")):
                r["ORT02"] = district
    return rows_by_sheet
