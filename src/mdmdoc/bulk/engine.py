#!/usr/bin/env python3
"""bulk.engine — one entry point for a bulk validation run.

Flow: read (template or raw export) -> case checks -> two outputs:
  * inbox/<bulk_id>__validated.xlsx — the operator's data COPIED with three
    appended columns (Verdict | Rule IDs | Reasons) + a Summary sheet. FULL
    values by design (local instance, the operator uploaded them — same
    posture as tin_bulk/addrval); inbox/ is gitignored.
  * runs/<bulk_id>/bulk_report.{json,md} — masked, leak-gated artifacts (row
    reasons are constructed mask-only in the case modules).
"""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

from .. import config, runstore
from . import reader
from .model import BUCKETS, BulkResult

_FILL = {"VALID": "C6EFCE", "SUSPICIOUS": "FFEB9C",
         "INVALID": "FFC7CE", "SKIPPED": "D9D9D9"}


def bulk_id_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def run_bulk(path: Path, case: str = "auto", refs: list | None = None,
             web: bool = False, progress=None) -> tuple[BulkResult, dict]:
    """-> (BulkResult, artifacts{result_xlsx, report_json, report_md, bulk_id}).
    progress: optional callable(str) for live job logs."""
    path = Path(path)
    _raw_say = progress or (lambda s: None)

    def say(msg: str) -> None:
        # cooperative cancel between row batches / web lookups (no-op when no
        # RunControl is active — CLI and tests run without one)
        from .. import runctl
        runctl.checkpoint("bulk")
        _raw_say(msg)
    if case in ("auto", "", None):
        case = reader.detect_case(path)
        say(f"detected case: {case}")
    if case not in ("bank", "tax", "region"):
        raise reader.BulkInputError(f"unknown bulk case '{case}'")

    rows, columns, kind = reader.read_rows(path, case)
    say(f"parsed {len(rows)} data rows ({kind}; columns: {', '.join(columns)})")
    res = BulkResult(case=case, source_file=path.name, source_kind=kind,
                     columns=columns, options={"web": bool(web)})

    if case == "tax":
        from . import tax
        res.rows = tax.check_rows(rows, progress=say)
    elif case == "bank":
        from . import bank
        res.rows = bank.check_rows(rows, web=web, notes=res.notes, progress=say)
    else:
        from . import region
        res.rows = region.check_rows(rows, refs=refs or [], notes=res.notes,
                                     progress=say)

    bulk_id = bulk_id_of(path)
    arts = {"bulk_id": bulk_id}
    arts["result_xlsx"] = str(_write_result_xlsx(path, case, rows, res, bulk_id))
    say(f"result workbook: {arts['result_xlsx']}")
    arts.update(_write_reports(res, bulk_id))
    return res, arts


# --- outputs ---------------------------------------------------------------------
def _write_result_xlsx(src: Path, case: str, rows: list[dict],
                       res: BulkResult, bulk_id: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    config.ensure_dirs()
    out = config.INBOX_DIR / f"{bulk_id}__{case}_validated.xlsx"
    wb = Workbook()

    from .reader import _CASE_ALIASES
    disp = {canon: names[0] for canon, names in _CASE_ALIASES[case].items()}
    ws = wb.active
    ws.title = "Data"
    headers = [disp.get(c, c) for c in res.columns] + ["Verdict", "Rule IDs", "Reasons"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row, rv in zip(rows, res.rows):
        ws.append([row.get(c, "") for c in res.columns]
                  + [rv.bucket, ", ".join(rv.rule_ids), " | ".join(rv.reasons)])
        fill = _FILL.get(rv.bucket)
        if fill:
            ws.cell(row=ws.max_row, column=len(res.columns) + 1).fill = \
                PatternFill("solid", fgColor=fill)
    ws.freeze_panes = "A2"

    s = wb.create_sheet("Summary")
    s.append(["Bulk validation summary"])
    s["A1"].font = Font(bold=True, size=14)
    s.append([])
    s.append(["Source file", res.source_file])
    s.append(["Recognized as", f"{case} ({res.source_kind})"])
    s.append(["Rows", len(res.rows)])
    s.append(["Options", json.dumps(res.options)])
    s.append([])
    s.append(["Bucket", "Rows"])
    for b in BUCKETS:
        s.append([b, res.counts().get(b, 0)])
    s.append([])
    s.append(["Rule", "Rows hit"])
    for rid, n in res.top_reasons():
        s.append([rid, n])
    for note in res.notes:
        s.append([])
        s.append(["Note", note])
    wb.save(out)
    return out


def _write_reports(res: BulkResult, bulk_id: str) -> dict:
    """Masked runs/ artifacts through the leak gate. Reasons are mask-only by
    construction (case modules); the strict pattern gate is the backstop."""
    summary = res.summary()
    problem_rows = [{"row": r.row_no, "partner": r.key, "bucket": r.bucket,
                     "rules": r.rule_ids, "reasons": r.reasons}
                    for r in res.rows if r.bucket != "VALID"][:2000]
    report = {"schema": "mdmdoc.bulk.v1", "bulk_id": bulk_id,
              "summary": summary, "problem_rows": problem_rows}
    runstore.write(bulk_id, "bulk_report.json", report, [], policy="strict")

    lines = [f"BULK VALIDATION — {res.case}", "",
             f"Source: {res.source_file} ({res.source_kind})",
             f"Rows: {len(res.rows)}", ""]
    for b in BUCKETS:
        lines.append(f"{b:<10} {res.counts().get(b, 0)}")
    lines.append("")
    lines.append("Top rules:")
    for rid, n in res.top_reasons():
        lines.append(f"  {rid}  — {n} row(s)")
    for note in res.notes:
        lines.append(f"Note: {note}")
    lines.append("")
    lines.append("First problem rows:")
    for r in problem_rows[:40]:
        lines.append(f"  row {r['row']} (BP {r['partner']}): {r['bucket']} "
                     f"[{', '.join(r['rules'])}] {'; '.join(r['reasons'])[:200]}")
    runstore.write(bulk_id, "bulk_report.md", "\n".join(lines) + "\n", [],
                   policy="strict")
    runstore.write(bulk_id, "meta.json",
                   {"run_id": bulk_id, "kind": "bulk", "case": res.case,
                    "file_name": res.source_file, "ts": runstore.now_iso(),
                    "rows": len(res.rows), "counts": res.counts(),
                    "options": res.options}, [], policy="strict")
    return {"report_json": f"runs/{bulk_id}/bulk_report.json",
            "report_md": f"runs/{bulk_id}/bulk_report.md"}
