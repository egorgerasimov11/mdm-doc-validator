#!/usr/bin/env python3
"""bulk.reader — one strict reader for BOTH input shapes:

  * the canonical templates (templates/bulk/*.xlsx — headers below verbatim);
  * raw SE16N exports whose descriptive headers alias to the same canon
    (the four real exports this feature was built against are the first
    aliases: BUT0BK bank dump, BP tax-number dump, T005S/T005U references).

Same discipline as sap_tables.load(): find the header row (first of the top 8
rows resolving >=2 known columns), map by normalized header, rows below become
dicts keyed by CANONICAL names. Unknown columns are ignored, never fatal.
"""
from __future__ import annotations

from ..sap_tables import _cell_str, _hkey


class BulkInputError(ValueError):
    pass


# --- canonical columns + aliases per case ---------------------------------------
# canon -> aliases (descriptive SE16N headers, template headers, technical names)
_TAX_ALIASES = {
    "partner": ("Business Partner", "Partner", "BP", "PARTNER"),
    "tax_category": ("Tax Number Category", "Tax Category", "TAXTYPE"),
    "tax_number": ("Tax Number", "Tax number", "TAXNUM"),
    "tax_number_long": ("Tax Number Long", "TAXNUMXL"),
    "country": ("Country", "Country/Region Key", "Country Key", "LAND1"),
    "name": ("Name", "Partner Name", "Name 1"),
}
_BANK_ALIASES = {
    "partner": ("Business Partner", "Partner", "BP", "PARTNER"),
    "bank_details_id": ("Bank Details ID", "BKVID"),
    "bank_country": ("Bank Country/Region", "Bank Country", "Bank Ctry", "BANKS"),
    "bank_key": ("Bank Key", "Bank Key (ABA Routing Code)", "Routing Number",
                 "ABA", "BANKL"),
    "bank_account": ("Bank Account", "Bank acct", "Bank Account Number",
                     "Account Number", "BANKN"),
    "iban": ("IBAN", "IBAN Number"),
    "swift_bic": ("SWIFT", "SWIFT/BIC", "BIC", "SWIFT Code", "BIC/Swift Code"),
    "control_key": ("Bank Control Key", "Control Key", "BKONT"),
    "account_holder": ("Account Holder", "Account Holder Name", "KOINH"),
    "account_name": ("Account Name", "ACCNAME"),
    "reference": ("Reference Details", "Reference", "BKREF"),
    "valid_from": ("Valid From",),
    "valid_to": ("Valid To",),
}
_REGION_ALIASES = {
    "partner": ("Business Partner", "Partner", "BP", "PARTNER"),
    "country": ("Country", "Country/Region Key", "Country Key", "Country/Region",
                "LAND1"),
    "region": ("Region", "State", "Province", "REGIO"),
    "postal_code": ("Postal Code", "Postal code", "ZIP", "ZIP Code", "PSTLZ"),
    "city": ("City", "ORT01"),
    "name": ("Name", "Partner Name", "Name 1"),
}

# marker canon columns that IDENTIFY a case when doc_case="auto"
_CASE_MARKERS = (
    ("tax", {"tax_category"}),
    ("bank", {"bank_key", "bank_account"}),
    ("region", {"region"}),           # after bank/tax — those may carry regions
)

_CASE_ALIASES = {"tax": _TAX_ALIASES, "bank": _BANK_ALIASES,
                 "region": _REGION_ALIASES}
# minimum canon columns a sheet must resolve to count as this case's input
_CASE_REQUIRED = {"tax": {"tax_category", "tax_number"},
                  "bank": {"bank_key", "bank_account"},
                  "region": {"country"}}


def _alias_map(case: str) -> dict:
    m: dict[str, str] = {}
    for canon, names in _CASE_ALIASES[case].items():
        m[_hkey(canon)] = canon
        for n in names:
            m[_hkey(n)] = canon
    return m


def _sheet_rows(path):
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise BulkInputError("openpyxl is required for bulk validation") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        yield from rows
    finally:
        wb.close()


def _resolve_header(row, amap: dict) -> dict:
    m: dict[str, int] = {}
    for c, cell in enumerate(row or ()):
        canon = amap.get(_hkey(cell))
        if canon and canon not in m:
            m[canon] = c
    return m


def detect_case(path) -> str:
    """auto-detect which case a workbook belongs to, by marker columns."""
    for i, row in enumerate(_sheet_rows(path)):
        if i >= 8:
            break
        for case, markers in _CASE_MARKERS:
            cols = set(_resolve_header(row, _alias_map(case)))
            if markers <= cols and _CASE_REQUIRED[case] <= cols:
                return case
    raise BulkInputError(
        "could not recognize the workbook as a banking / tax / postal-region "
        "input — download the matching template from the Bulk page and use "
        "its Data-sheet headers")


def read_rows(path, case: str) -> tuple[list[dict], list[str], str]:
    """-> (rows as canonical dicts, canonical columns seen, source_kind).
    source_kind: 'template' when every resolved header matched the canon name
    itself, else 'raw-export' (aliased descriptive headers)."""
    amap = _alias_map(case)
    header_idx, col_map = None, {}
    header_row = None
    for i, row in enumerate(_sheet_rows(path)):
        if i >= 8:
            break
        m = _resolve_header(row, amap)
        if len(m) >= 2 and _CASE_REQUIRED[case] <= set(m):
            header_idx, col_map, header_row = i, m, row
            break
    if header_idx is None:
        need = ", ".join(sorted(_CASE_REQUIRED[case]))
        raise BulkInputError(
            f"no recognizable {case} header row found (need at least: {need}) "
            f"— use the {case} template's Data sheet")
    canon_hdrs = {_hkey(c) for c in _CASE_ALIASES[case]}
    matched = [_hkey(header_row[i]) for i in col_map.values()
               if i < len(header_row or ())]
    kind = "template" if all(h in canon_hdrs for h in matched) else "raw-export"
    rows: list[dict] = []
    for i, row in enumerate(_sheet_rows(path)):
        if i <= header_idx:
            continue
        d = {canon: _cell_str(row[c]) if c < len(row) else ""
             for canon, c in col_map.items()}
        if any(d.values()):
            rows.append(d)
    return rows, sorted(col_map), kind


# --- reference tables (attached per run) -----------------------------------------
def read_t005s(path) -> dict:
    """T005S region master -> {ISO2: {region_code: description}}. Recognizes the
    SE16N export shape: Country/Region Key | Region | ... | Description."""
    amap = {_hkey(h): c for h, c in (
        ("Country/Region Key", "country"), ("Country Key", "country"),
        ("Country", "country"), ("Region", "region"),
        ("Description", "description"))}
    header_idx, col_map = None, {}
    for i, row in enumerate(_sheet_rows(path)):
        if i >= 8:
            break
        m: dict[str, int] = {}
        for c, cell in enumerate(row or ()):
            canon = amap.get(_hkey(cell))
            if canon and canon not in m:
                m[canon] = c
        if {"country", "region"} <= set(m):
            header_idx, col_map = i, m
            break
    if header_idx is None:
        raise BulkInputError("not a T005S region export (no Country/Region headers)")
    out: dict[str, dict] = {}
    for i, row in enumerate(_sheet_rows(path)):
        if i <= header_idx:
            continue
        cc = _cell_str(row[col_map["country"]]) if col_map["country"] < len(row) else ""
        rg = _cell_str(row[col_map["region"]]) if col_map["region"] < len(row) else ""
        de = (_cell_str(row[col_map["description"]])
              if "description" in col_map and col_map["description"] < len(row) else "")
        if cc and rg:
            out.setdefault(cc.upper(), {})[rg] = de
    return out


def read_t005u(path) -> dict:
    """T005U region TEXTS (language-keyed) -> {(lang, ISO2): {region: text}}.
    Optional — only used to annotate region descriptions in the result."""
    amap = {_hkey(h): c for h, c in (
        ("Language Key", "lang"), ("Country/Region Key", "country"),
        ("Country Key", "country"), ("Region", "region"),
        ("Description", "description"))}
    header_idx, col_map = None, {}
    for i, row in enumerate(_sheet_rows(path)):
        if i >= 8:
            break
        m: dict[str, int] = {}
        for c, cell in enumerate(row or ()):
            canon = amap.get(_hkey(cell))
            if canon and canon not in m:
                m[canon] = c
        if {"lang", "country", "region"} <= set(m):
            header_idx, col_map = i, m
            break
    if header_idx is None:
        raise BulkInputError("not a T005U region-text export")
    out: dict = {}
    for i, row in enumerate(_sheet_rows(path)):
        if i <= header_idx:
            continue
        vals = {k: (_cell_str(row[c]) if c < len(row) else "")
                for k, c in col_map.items()}
        if vals.get("country") and vals.get("region"):
            out.setdefault((vals.get("lang", ""), vals["country"].upper()),
                           {})[vals["region"]] = vals.get("description", "")
    return out
