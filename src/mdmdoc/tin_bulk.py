#!/usr/bin/env python3
"""tin_bulk.py — bulk US TIN validation for spreadsheet exports (the /ui/tax tab).

Port of the us-tax-number-validator skill (Vault_tools/Agent/skills/
us-tax-number-validator; rules re-verified against IRS/SSA sources 2026-07-10).
Buckets every row into Valid / Suspicious / Not valid / Dummy-Masked judging
the number's own FORMAT & STRUCTURE — never the SAP tax category (there is no
authoritative US0-US4 -> SSN/EIN mapping). Placeholder policy per Egor
2026-07-09: punctuation-only and all-zeros values are placeholders -> Dummy.

The IRS/SSA rule tables live in rules/predicates.py and are shared with the
W-9 document rules W9-040/W9-041 — one source of truth on this side of the
validator. Output workbook carries full TINs by design (local instance, same
posture as the /address tab); HTML surfaces only counts and masked samples.
"""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from .rules.predicates import (_EIN_FMT, _KNOWN_FAKE_TINS, _SSN_FMT, _ein_bad,
                               _itin_bad, _ssn_bad, _tin_placeholder_kind)

VALID, SUSPECT, NOTAX, DUMMY = "Valid", "Suspicious", "Not valid", "Dummy / Masked"
BUCKETS = (VALID, SUSPECT, NOTAX, DUMMY)

_DOM_WORD = re.compile(r"^(FEIN|EIN|SSN|TIN)\b[\s:.\-]*(.+)$", re.I)
_CC_PAIR = re.compile(r"^([A-Za-z]{2})[\s\-]?(\d[\d\s\-]*)$")
_RAW9 = re.compile(r"^\d{9}$")
_PUNCT = re.compile(r"^[.\-,/ ]+$")


def _digits(s: str) -> str:
    return re.sub(r"\D", "", s)


def classify(raw) -> tuple[str, str, str]:
    """-> (verdict, detected_type, reason) for one raw cell value."""
    s = ("" if raw is None else str(raw)).strip()
    if s == "":
        return NOTAX, "", "blank / empty"
    if re.fullmatch(r"[Xx]+", s):
        return DUMMY, "masked", f"masked value ({s})"
    if _PUNCT.fullmatch(s):
        return DUMMY, "placeholder", f"placeholder / empty (only punctuation: {s})"

    had_prefix, core = None, s
    m = _DOM_WORD.match(s)
    if m:
        had_prefix, core = m.group(1).upper(), m.group(2).strip()
    else:
        mp = _CC_PAIR.match(s)
        if mp:
            cc = mp.group(1).upper()
            if cc == "US":
                had_prefix, core = "US", mp.group(2).strip()
            else:
                return NOTAX, f"{cc}?", f'non-US tax ID / VAT (prefix "{cc}")'

    d = _digits(core)
    if d and set(d) == {"0"}:
        return DUMMY, "placeholder", f"all-zeros placeholder ({s})"
    if len(d) >= 3 and len(set(d)) == 1:
        kind = "all-nines" if d[0] == "9" else f"repeated {d[0]}"
        return DUMMY, "placeholder", f"{kind} placeholder ({s})"
    if len(d) >= 10 and set(d) <= {"0", "9"}:
        return DUMMY, "placeholder", f"placeholder (only 0s/9s, {len(d)} digits: {s})"

    if re.search(r"[A-Za-z]", core):
        return NOTAX, "", "contains letters (not a numeric TIN)"
    if not d:
        return NOTAX, "", "no digits"
    if len(d) <= 6:
        return NOTAX, "", f"too short ({len(d)} digits)"
    if len(d) >= 12:
        return NOTAX, "", f"too long for a US TIN ({len(d)} digits)"
    if len(d) != 9:
        r = f"wrong length ({len(d)} digits; a US TIN has 9)"
        if had_prefix:
            r += f'; also had "{had_prefix}" prefix'
        return SUSPECT, "", r
    if d in _KNOWN_FAKE_TINS:
        return DUMMY, "placeholder", f"known never-issued / reserved fake TIN ({s})"
    if _tin_placeholder_kind(d):
        return DUMMY, "placeholder", f"placeholder ({s})"

    verdict, dtype, reason = SUSPECT, "", ""
    if _EIN_FMT.match(core):
        dtype = "EIN"
        bad = _ein_bad(d)
        if bad:
            verdict, reason = SUSPECT, f'EIN format but "{d[:2]}" is a never-assigned IRS prefix'
        else:
            verdict, reason = VALID, "EIN"
    elif _SSN_FMT.match(core):
        if d[0] == "9":
            dtype = "ITIN"
            bad = _itin_bad(d)
            verdict, reason = (SUSPECT, f"{bad} (group {d[3:5]})") if bad else (VALID, "ITIN")
        else:
            dtype = "SSN"
            bad = _ssn_bad(d)
            verdict, reason = (SUSPECT, bad) if bad else (VALID, "SSN")
    elif _RAW9.match(core):
        oks = []
        if not _ein_bad(d):
            oks.append("EIN")
        if d[0] != "9" and not _ssn_bad(d):
            oks.append("SSN")
        if d[0] == "9" and not _itin_bad(d):
            oks.append("ITIN")
        if oks:
            dtype = "/".join(oks)
            verdict, reason = VALID, f"9 digits, no separator; valid as {', '.join(oks)}"
        else:
            verdict, reason = SUSPECT, "9 digits but fails EIN prefix, SSN and ITIN structure"
    else:
        verdict, reason = SUSPECT, f"nonstandard grouping for a 9-digit value ({core})"

    if had_prefix:
        if verdict == VALID:
            verdict = SUSPECT
            reason = f'valid {dtype} but field contains extra "{had_prefix}" prefix'
        else:
            reason += f'; also had "{had_prefix}" prefix'
    return verdict, dtype, reason


# ---- column resolution (SAP export header aliases) -----------------------------
_HEADER_ALIASES = {
    "tax": ["tax number", "tax_number", "taxnumber", "tax no", "tax id",
            "steuernummer", "stcd1", "stcd2", "vat"],
    "id": ["business partner", "bp", "partner", "vendor", "customer",
           "lifnr", "kunnr", "account"],
    "cat": ["tax number category", "category", "tax category", "tax type"],
}


def _resolve_col(spec, headers):
    if spec is None or str(spec).strip() == "":
        return None
    spec = str(spec).strip()
    low = [(h or "").strip().lower() for h in headers]
    if spec.lower() in low:
        return low.index(spec.lower())
    if re.fullmatch(r"[A-Za-z]{1,3}", spec):
        try:
            return column_index_from_string(spec.upper()) - 1
        except ValueError:
            pass
    if spec.isdigit():
        return int(spec) - 1
    return None


def _auto_col(kind, headers):
    low = [(h or "").strip().lower() for h in headers]
    for alias in _HEADER_ALIASES[kind]:
        if alias in low:
            return low.index(alias)
    for alias in _HEADER_ALIASES[kind]:
        for i, h in enumerate(low):
            if alias in h:
                return i
    return None


def validate_workbook(src: Path, out: Path, sheet: str | None = None,
                      col=None, id_col=None, cat_col=None) -> dict:
    """Classify every row of `src` and write the bucketed workbook to `out`.
    -> summary dict: counts per bucket, top reasons, resolved column labels."""
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("empty sheet")
    headers = list(rows[0])
    data = rows[1:]

    tax_i = _resolve_col(col, headers)
    if tax_i is None:
        tax_i = _auto_col("tax", headers)
    if tax_i is None:
        raise ValueError(f"tax-number column not found; headers: {headers}")
    id_i = _resolve_col(id_col, headers)
    if id_i is None:
        id_i = _auto_col("id", headers)
    cat_i = _resolve_col(cat_col, headers)
    if cat_i is None:
        cat_i = _auto_col("cat", headers)

    buckets = {b: [] for b in BUCKETS}
    reasons = Counter()
    types = Counter()
    for n, r in enumerate(data, start=2):
        def cell(i):
            return "" if (i is None or i >= len(r) or r[i] is None) else str(r[i])
        num = cell(tax_i)
        verdict, dtype, reason = classify(num)
        buckets[verdict].append([cell(id_i) if id_i is not None else str(n),
                                 cell(cat_i), num, dtype, verdict, reason])
        reasons[f"[{verdict}] " + reason.split(";")[0].split(" (")[0]] += 1
        if verdict == VALID:
            types[dtype] += 1

    labels = (headers[id_i] if id_i is not None else "Row",
              headers[cat_i] if cat_i is not None else "Category",
              headers[tax_i])
    _write_workbook(buckets, reasons, types, labels, out)
    total = sum(len(v) for v in buckets.values())
    return {"total": total,
            "counts": {b: len(buckets[b]) for b in BUCKETS},
            "top_reasons": reasons.most_common(12),
            "valid_types": types.most_common(),
            "columns": {"tax": str(labels[2]), "id": str(labels[0]),
                        "category": str(labels[1])},
            "out": str(out)}


def _write_workbook(buckets, reasons, types, labels, out: Path) -> None:
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    base = Font(name="Arial")
    fills = {VALID: "E2EFDA", SUSPECT: "FFF2CC", NOTAX: "FCE4D6", DUMMY: "D9D9D9"}
    sheet_names = {VALID: "Valid", SUSPECT: "Suspicious",
                   NOTAX: "Not valid", DUMMY: "Dummy-Masked"}
    headers = [labels[0], labels[1], labels[2], "Detected type", "Verdict", "Reason"]

    wb = Workbook()
    wb.remove(wb.active)
    for bucket in BUCKETS:
        sh = wb.create_sheet(sheet_names[bucket])
        sh.append(headers)
        for c in range(1, len(headers) + 1):
            cc = sh.cell(1, c)
            cc.font, cc.fill = hdr_font, hdr_fill
        for row in buckets[bucket]:
            sh.append(row)
        for i, w in enumerate((20, 18, 22, 14, 16, 62), 1):
            sh.column_dimensions[get_column_letter(i)].width = w
        sh.freeze_panes = "A2"
        if buckets[bucket]:
            sh.auto_filter.ref = f"A1:F{len(buckets[bucket]) + 1}"

    total = sum(len(v) for v in buckets.values())
    sm = wb.create_sheet("Summary", 0)
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["A"].width = 46
    sm.column_dimensions["B"].width = 12
    sm["A1"] = "US Tax Number Validation — Summary"
    sm["A1"].font = Font(name="Arial", bold=True, size=14)
    sm["A2"] = f"{total} rows analyzed (mdmdoc /tax; rules: IRS/SSA via W9-040/041 tables)"
    sm["A2"].font = base
    row = 4
    sm[f"A{row}"] = "Verdict"
    sm[f"B{row}"] = "Count"
    ref = {VALID: "Valid", SUSPECT: "Suspicious",
           NOTAX: "'Not valid'", DUMMY: "'Dummy-Masked'"}
    row += 1
    first = row
    for b in BUCKETS:
        sm[f"A{row}"] = b
        sm[f"A{row}"].fill = PatternFill("solid", fgColor=fills[b])
        sm[f"B{row}"] = f"=COUNTA({ref[b]}!A2:A1048576)"
        row += 1
    sm[f"A{row}"] = f"TOTAL (must equal {total})"
    sm[f"B{row}"] = f"=SUM(B{first}:B{row - 1})"
    sm[f"A{row + 1}"] = "Integrity check"
    sm[f"B{row + 1}"] = f'=IF(B{row}={total},"OK","MISMATCH")'
    sm[f"B{row + 1}"].font = Font(name="Arial", bold=True)
    row += 3
    sm[f"A{row}"] = "Valid — by detected type"
    row += 1
    for t, cnt in types.most_common():
        sm[f"A{row}"] = f"   {t}"
        sm[f"B{row}"] = cnt
        row += 1
    row += 1
    sm[f"A{row}"] = "Top reason codes"
    row += 1
    for reason, cnt in reasons.most_common(20):
        sm[f"A{row}"] = reason[:90]
        sm[f"B{row}"] = cnt
        row += 1
    row += 1
    for line in ("Judged by FORMAT/STRUCTURE only — not by SAP category.",
                 "'Valid' = well-formed & plausible, NOT 'confirmed to exist'",
                 "(IRS TIN Matching / SSA CBSV are gated; no open per-number lookup)."):
        sm[f"A{row}"] = line
        sm[f"A{row}"].font = Font(name="Arial", italic=True, size=9)
        row += 1
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
